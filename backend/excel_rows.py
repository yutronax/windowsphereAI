import os
import shutil
import tempfile
from pathlib import Path

import openpyxl


def _normalize_rows(rows: list) -> list[list]:
    """Saga #326 (ATDD AC-3): LLM/kullanıcı düz bir liste (`[1, 2, 3]`)
    gönderebilir - iç içe (`[[1], [2], [3]]`) yerine. Her eleman zaten
    list/tuple ise olduğu gibi bırakılır, değilse tek-hücreli bir satıra
    sarılır. `create_excel_file`/`append_excel_rows` arasında paylaşılan
    ortak normalizasyon (ATDD: "ikisinde de aynı sarma mantığı")."""
    normalized = []
    for row in rows:
        if isinstance(row, (list, tuple)):
            normalized.append(list(row))
        else:
            normalized.append([row])
    return normalized


def create_excel_file(rows: list, destination_path: Path) -> None:
    """Saga #326 (ATDD AC-1/AC-2/AC-3): `rows`ı yeni bir `.xlsx` dosyasına
    (`destination_path`) yazar. Hedef ZATEN VARSA `FileExistsError`
    fırlatılır, hiçbir şey yazılmaz/değiştirilmez (AC-2). `excel_sort.py`
    ile AYNI geçici-dosya-yaz + atomik `Path.replace` deseni kullanılır."""
    if destination_path.exists():
        raise FileExistsError(
            f"EXCEL_CREATE çıktı dosyası zaten var: '{destination_path.name}'"
        )

    normalized_rows = _normalize_rows(rows)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in normalized_rows:
        sheet.append(row)

    fd, temp_path_str = tempfile.mkstemp(
        suffix=".tmp", prefix=destination_path.name + ".", dir=str(destination_path.parent)
    )
    temp_path = Path(temp_path_str)
    os.close(fd)
    try:
        workbook.save(str(temp_path))
        temp_path.replace(destination_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def append_excel_rows(source_path: Path, rows: list, backup_path: Path) -> None:
    """Saga #326 (ATDD AC-6/AC-7): `rows`ı `source_path`in SONUNA ekler,
    kaynağı YERİNDE günceller (PDF `_forward_append`'in AYNI "önce oku,
    sonra yedekle, sonra atomik yaz" sırası - openpyxl'in doğrudan
    `wb.save(source_path)` kolaylığına GÜVENİLMEZ, plan.md'nin Risks
    notu). Kaynak yoksa/bozuksa exception fırlatılır, HİÇBİR dosya
    oluşturulmaz/değiştirilmez (AC-7)."""
    if not source_path.exists():
        raise FileNotFoundError(
            f"EXCEL_APPEND kaynağı okunamıyor: '{source_path.name}'"
        )

    try:
        workbook = openpyxl.load_workbook(str(source_path))
    except Exception as exc:
        raise ValueError(
            f"EXCEL_APPEND kaynağı okunamıyor: '{source_path.name}' ({exc})"
        ) from exc

    sheet = workbook.active
    normalized_rows = _normalize_rows(rows)
    for row in normalized_rows:
        sheet.append(row)

    backup_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(source_path), str(backup_path))

    fd, temp_path_str = tempfile.mkstemp(
        suffix=".tmp", prefix=source_path.name + ".", dir=str(source_path.parent)
    )
    temp_path = Path(temp_path_str)
    os.close(fd)
    try:
        workbook.save(str(temp_path))
        temp_path.replace(source_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def read_excel_range(source_path: Path, range_spec: str | None) -> list[list]:
    """Saga #326 (ATDD AC-4/AC-5): `source_path`teki `.xlsx` dosyasının
    aktif sayfasını okur. `range_spec` verilmişse ("A1:C10") SADECE o
    hücre aralığı döner (AC-4); verilmemişse tüm kullanılan alan döner
    (AC-5, `ws.iter_rows()`). Geçersiz `range_spec` openpyxl'in kendi
    `ValueError`'ını yansıtır - buraya özel bir try/except EKLENMEZ."""
    workbook = openpyxl.load_workbook(str(source_path), data_only=True)
    sheet = workbook.active

    if range_spec is None:
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    cells = sheet[range_spec]
    if isinstance(cells, tuple):
        if cells and not isinstance(cells[0], tuple):
            # Tek bir satır (ör. "A1:C1") - openpyxl bir Cell tuple'ı
            # döndürür (tuple-of-tuples değil), tek satırlık bir sonuca
            # normalize edilir.
            return [[cell.value for cell in cells]]
        return [[cell.value for cell in row] for row in cells]

    # Tek bir hücre (ör. "A1") - openpyxl tek bir Cell nesnesi döndürür.
    return [[cells.value]]
