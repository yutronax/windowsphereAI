# Saga #325: excel-satir-filtreleme (red step) - backend/excel_sort.py'ye
# `filter_excel_sheet` / `ExcelFilterFormulaGuardError` henuz eklenmedi, bu
# dosyanin TAMAMI import asamasinda basarisiz olmali (ImportError) -
# beklenen kirmizi durum. code-copilot bu isimleri excel_sort.py'ye
# ekledikten sonra bu testler calisir hale gelmeli.
#
# `filter_excel_sheet(source_path: Path, filter_column: str, filter_value,
# destination_path: Path) -> None` ve `ExcelFilterFormulaGuardError`,
# atdd.md/plan.md'de tasarlanan isimlerdir (bkz.
# artifacts/excel-satir-filtreleme/plan.md, Davranış Sözleşmesi tablosu).
# Sütun çözümlemesi test_excel_sort.py'deki resolve_sort_column ile AYNI
# fonksiyon üzerinden yapılır (header-önce/harf-sonra) - burada ayrıca
# test edilmiyor, sadece filter_excel_sheet üzerinden entegrasyon olarak
# doğrulanıyor.

import openpyxl
import pytest

from backend.excel_sort import ExcelFilterFormulaGuardError, filter_excel_sheet


def _write_workbook(path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(str(path))


def _write_workbook_with_formula(path, header: list, data_rows: list[list], formula_cell: str, formula: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in data_rows:
        ws.append(row)
    ws[formula_cell] = formula
    wb.save(str(path))


# --- filter_excel_sheet: header-first column resolution (AC-1, AC-2) ---


def test_filter_excel_sheet_matches_by_header_text_and_keeps_matching_rows(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "filtered.xlsx"
    _write_workbook(
        source,
        [
            ["Ad", "Puan"],
            ["Ali", 90],
            ["Veli", 80],
            ["Can", 90],
        ],
    )

    filter_excel_sheet(source, "Puan", 90, destination)

    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows == [("Ad", "Puan"), ("Ali", 90), ("Can", 90)]


def test_filter_excel_sheet_prefers_header_text_over_bare_letter_interpretation(tmp_path):
    # "Ad" gibi kisa/alfabetik bir baslik sutun harfi ("AD") ile
    # karistirilmamali (Saga #325 P0 senaryosu, ATDD AC-2).
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "filtered.xlsx"
    _write_workbook(
        source,
        [
            ["Ad", "Puan"],
            ["Ali", 90],
            ["Veli", 80],
        ],
    )

    filter_excel_sheet(source, "Ad", "Ali", destination)

    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows == [("Ad", "Puan"), ("Ali", 90)]


def test_filter_excel_sheet_supports_bare_column_letter_when_no_header_matches(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "filtered.xlsx"
    _write_workbook(
        source,
        [
            ["Ad", "Puan"],
            ["Ali", 90],
            ["Veli", 80],
        ],
    )

    filter_excel_sheet(source, "B", 90, destination)

    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows == [("Ad", "Puan"), ("Ali", 90)]


def test_filter_excel_sheet_raises_value_error_for_unknown_column(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "filtered.xlsx"
    _write_workbook(source, [["Ad", "Puan"], ["Ali", 90]])

    with pytest.raises(ValueError):
        filter_excel_sheet(source, "BilinmeyenSutun", 90, destination)

    assert not destination.exists()


# --- filter_excel_sheet: source is never modified (Davranış Sözleşmesi #1) ---


def test_filter_excel_sheet_does_not_modify_the_source_file(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "filtered.xlsx"
    _write_workbook(source, [["Ad", "Puan"], ["Ali", 90], ["Veli", 80]])
    original_bytes = source.read_bytes()

    filter_excel_sheet(source, "Puan", 90, destination)

    assert source.read_bytes() == original_bytes


# --- filter_excel_sheet: zero matches is success, not silent no-op (AC-4) ---


def test_filter_excel_sheet_writes_header_only_file_when_no_row_matches(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "filtered.xlsx"
    _write_workbook(source, [["Ad", "Puan"], ["Ali", 90], ["Veli", 80]])

    filter_excel_sheet(source, "Puan", 999, destination)

    assert destination.exists()
    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows == [("Ad", "Puan")]


# --- filter_excel_sheet: formula guard (AC-5) ---


def test_filter_excel_sheet_raises_formula_guard_error_when_a_data_cell_is_a_real_formula(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "filtered.xlsx"
    _write_workbook_with_formula(
        source,
        header=["Ad", "Puan"],
        data_rows=[["Ali", 90], ["Veli", 80]],
        formula_cell="C2",
        formula="=B2+B3",
    )

    with pytest.raises(ExcelFilterFormulaGuardError):
        filter_excel_sheet(source, "Puan", 90, destination)

    assert not destination.exists()


def test_filter_excel_sheet_leaves_source_untouched_when_formula_guard_triggers(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "filtered.xlsx"
    _write_workbook_with_formula(
        source,
        header=["Ad", "Puan"],
        data_rows=[["Ali", 90], ["Veli", 80]],
        formula_cell="C2",
        formula="=B2+B3",
    )
    original_bytes = source.read_bytes()

    with pytest.raises(ExcelFilterFormulaGuardError):
        filter_excel_sheet(source, "Puan", 90, destination)

    assert source.read_bytes() == original_bytes


# --- filter_excel_sheet: empty data no-op (AC-6) ---


def test_filter_excel_sheet_is_a_no_op_copy_when_only_header_row_exists(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "filtered.xlsx"
    _write_workbook(source, [["Ad", "Puan"]])

    filter_excel_sheet(source, "Puan", 90, destination)

    assert destination.exists()
    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows == [("Ad", "Puan")]


def test_filter_excel_sheet_is_a_no_op_copy_for_a_completely_empty_sheet(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "filtered.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(source))

    # Bos sayfada bile bir sutun spesifigi verilebilir; no-op durumda
    # sutun cozumlemesi/formul taramasi ANLAMSIZ oldugu icin hata
    # FIRLATILMAMALI (test_sort_excel_sheet_is_a_no_op_copy_for_a_
    # completely_empty_sheet ile AYNI sozlesme).
    filter_excel_sheet(source, "A", 90, destination)

    assert destination.exists()
