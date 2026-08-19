# Saga #326: excel-create-read-append (TEST-FIRST / red step) -
# `backend/excel_rows.py` (create_excel_file/append_excel_rows/
# read_excel_range) henuz YOK. Bu testler simdilik KIRMIZI kalmali
# (ModuleNotFoundError/ImportError - beklenen red durumu).
# Referans: artifacts/excel-create-read-append/atdd.md (AC-1, AC-3, AC-4,
# AC-5, AC-6, AC-7), plan.md (create_excel_file/append_excel_rows/
# read_excel_range imzalari, backup-once-onus deseni).

import openpyxl
import pytest

from backend.excel_rows import (
    append_excel_rows,
    create_excel_file,
    read_excel_range,
)


# ---------------------------------------------------------------------------
# create_excel_file(rows, destination_path)
# ---------------------------------------------------------------------------


def test_create_excel_file_writes_the_given_rows(tmp_path):
    # AC-1: verilen satirlarla dosya tam olarak oluşur.
    destination = tmp_path / "yeni.xlsx"
    rows = [["Ad", "Puan"], ["Ali", 90]]

    create_excel_file(rows, destination)

    assert destination.exists()
    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    values = [list(row) for row in ws.iter_rows(values_only=True)]
    assert values == rows


def test_create_excel_file_raises_when_destination_already_exists_and_does_not_overwrite(tmp_path):
    # AC-2: hedef zaten VARSA dosyaya DOKUNULMAZ (üzerine yazılmaz).
    destination = tmp_path / "var-olan.xlsx"
    original_rows = [["Eski", "Veri"]]
    create_excel_file_setup_wb = openpyxl.Workbook()
    ws = create_excel_file_setup_wb.active
    for row in original_rows:
        ws.append(row)
    create_excel_file_setup_wb.save(str(destination))
    original_bytes = destination.read_bytes()

    with pytest.raises(Exception):
        create_excel_file([["Yeni", "Veri"]], destination)

    assert destination.read_bytes() == original_bytes


def test_create_excel_file_wraps_a_flat_row_list_into_single_cell_rows(tmp_path):
    # AC-3: düz `rows=[1,2,3]` her eleman tek hücreli bir satır olarak
    # sarılır (`[[1],[2],[3]]`), çökme YOK.
    destination = tmp_path / "duz-liste.xlsx"

    create_excel_file([1, 2, 3], destination)

    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    values = [list(row) for row in ws.iter_rows(values_only=True)]
    assert values == [[1], [2], [3]]


# ---------------------------------------------------------------------------
# append_excel_rows(source_path, rows, backup_path)
# ---------------------------------------------------------------------------


def _write_excel(root, filename: str, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(str(root / filename))


def test_append_excel_rows_appends_new_rows_to_the_end_and_keeps_previous_content(tmp_path):
    # AC-6: satırlar dosyanın SONUNA eklenir, önceki içerik korunur (kaynak
    # YERİNDE güncellenir).
    source = tmp_path / "kaynak.xlsx"
    backup = tmp_path / "yedek" / "kaynak.xlsx"
    _write_excel(tmp_path, "kaynak.xlsx", [["Ad", "Puan"], ["Ali", 90]])

    append_excel_rows(source, [["Veli", 80]], backup)

    wb = openpyxl.load_workbook(str(source))
    ws = wb.active
    values = [list(row) for row in ws.iter_rows(values_only=True)]
    assert values == [["Ad", "Puan"], ["Ali", 90], ["Veli", 80]]


def test_append_excel_rows_raises_when_source_missing_and_creates_no_file(tmp_path):
    # AC-7: kaynak dosya YOK -> exception, HİÇBİR yeni/boş dosya oluşmaz.
    source = tmp_path / "yok.xlsx"
    backup = tmp_path / "yedek" / "yok.xlsx"

    with pytest.raises(Exception):
        append_excel_rows(source, [["Veli", 80]], backup)

    assert not source.exists()
    assert not backup.exists()


def test_append_excel_rows_raises_when_source_is_corrupt_and_leaves_it_untouched(tmp_path):
    # AC-7: kaynak BOZUKSA (geçersiz xlsx içeriği) -> exception, kaynak
    # değişmez.
    source = tmp_path / "bozuk.xlsx"
    corrupt_bytes = b"not a real xlsx"
    source.write_bytes(corrupt_bytes)
    backup = tmp_path / "yedek" / "bozuk.xlsx"

    with pytest.raises(Exception):
        append_excel_rows(source, [["Veli", 80]], backup)

    assert source.read_bytes() == corrupt_bytes


def test_append_excel_rows_backs_up_the_source_before_writing_the_new_content(tmp_path):
    # `backup_path`'e yazmadan ÖNCE kaynağın bir kopyasının alındığını
    # doğrula (backup dosyası eklemeden ÖNCEKİ içerikle eşleşmeli).
    source = tmp_path / "kaynak.xlsx"
    backup = tmp_path / "yedek" / "kaynak.xlsx"
    _write_excel(tmp_path, "kaynak.xlsx", [["Ad", "Puan"], ["Ali", 90]])
    pre_append_bytes = source.read_bytes()

    append_excel_rows(source, [["Veli", 80]], backup)

    assert backup.exists()
    assert backup.read_bytes() == pre_append_bytes
    wb = openpyxl.load_workbook(str(backup))
    ws = wb.active
    backup_values = [list(row) for row in ws.iter_rows(values_only=True)]
    assert backup_values == [["Ad", "Puan"], ["Ali", 90]]


# ---------------------------------------------------------------------------
# read_excel_range(source_path, range_spec)
# ---------------------------------------------------------------------------


def test_read_excel_range_returns_the_whole_used_area_when_range_spec_is_none(tmp_path):
    # AC-5: `range_spec=None` iken tüm kullanılan alan döner.
    source = tmp_path / "kaynak.xlsx"
    _write_excel(tmp_path, "kaynak.xlsx", [["Ad", "Puan"], ["Ali", 90], ["Veli", 80]])

    values = read_excel_range(source, None)

    assert values == [["Ad", "Puan"], ["Ali", 90], ["Veli", 80]]


def test_read_excel_range_returns_only_the_given_range_when_range_spec_is_provided(tmp_path):
    # AC-4: `range_spec="A1:B2"` iken SADECE o aralık döner (tüm sayfa
    # değil).
    source = tmp_path / "kaynak.xlsx"
    _write_excel(
        tmp_path,
        "kaynak.xlsx",
        [["Ad", "Puan", "Sehir"], ["Ali", 90, "Ankara"], ["Veli", 80, "Izmir"]],
    )

    values = read_excel_range(source, "A1:B2")

    assert values == [["Ad", "Puan"], ["Ali", 90]]


def test_read_excel_range_raises_value_error_for_an_invalid_range(tmp_path):
    # geçersiz range ("ZZZ9999999" gibi) `ValueError` fırlatır.
    source = tmp_path / "kaynak.xlsx"
    _write_excel(tmp_path, "kaynak.xlsx", [["Ad", "Puan"], ["Ali", 90]])

    with pytest.raises(ValueError):
        read_excel_range(source, "ZZZ9999999")
