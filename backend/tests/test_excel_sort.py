# Saga #324: excel-sort-formula-guard (red step) - backend/excel_sort.py
# henuz mevcut degil, bu dosyanin TAMAMI import asamasinda basarisiz
# olmali (ModuleNotFoundError / collection error) - beklenen kirmizi
# durum. Implementasyon subagent'i modulu yazdiktan sonra bu testler
# calisir hale gelmeli.
#
# `resolve_sort_column(header_row, column_spec) -> int` ve
# `sort_excel_sheet(source_path: Path, sort_column: str, ascending: bool,
# destination_path: Path) -> None` ile `ExcelSortFormulaGuardError`
# plan.md'de tasarlanan isimlerdir (bkz.
# artifacts/excel-sort-formula-guard/plan.md bolum 4).

import openpyxl
import pytest

from backend.excel_sort import (
    ExcelSortFormulaGuardError,
    resolve_sort_column,
    sort_excel_sheet,
)


def _write_workbook(path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(str(path))


def _write_workbook_with_formula(path, header: list, data_rows: list[list], formula_cell: str, formula: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in data_rows:
        ws.append(row)
    ws[formula_cell] = formula
    wb.save(str(path))


# --- resolve_sort_column ---


def test_resolve_sort_column_matches_header_text_case_insensitively():
    header_row = ("Ad", "Tutar", "Tarih")

    index = resolve_sort_column(header_row, "tutar")

    assert index == 1


def test_resolve_sort_column_matches_header_text_with_surrounding_whitespace():
    header_row = ("Ad", "Tutar", "Tarih")

    index = resolve_sort_column(header_row, "  Tutar  ")

    assert index == 1


def test_resolve_sort_column_falls_back_to_bare_column_letter_when_no_header_matches():
    header_row = ("Ad", "Tutar", "Tarih")

    index = resolve_sort_column(header_row, "B")

    assert index == 1


def test_resolve_sort_column_prefers_header_text_over_bare_letter_interpretation():
    # Header'da literal olarak "B" adinda bir sutun varsa, bare-letter
    # fallback'ten ONCE header eslesmesi denenmeli (ATDD P0 #3).
    header_row = ("A", "B", "C")

    index = resolve_sort_column(header_row, "B")

    assert index == 1


def test_resolve_sort_column_raises_value_error_for_unknown_column():
    header_row = ("Ad", "Tutar", "Tarih")

    with pytest.raises(ValueError):
        resolve_sort_column(header_row, "Bilinmeyen")


def test_resolve_sort_column_raises_value_error_for_blank_column_spec():
    header_row = ("Ad", "Tutar", "Tarih")

    with pytest.raises(ValueError):
        resolve_sort_column(header_row, "   ")


# --- sort_excel_sheet: formula guard ---


def test_sort_excel_sheet_raises_formula_guard_error_when_a_data_cell_is_a_real_formula(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "sorted.xlsx"
    _write_workbook_with_formula(
        source,
        header=["Ad", "Tutar"],
        data_rows=[["Ali", 100], ["Veli", 200]],
        formula_cell="C2",
        formula="=B2+B3",
    )

    with pytest.raises(ExcelSortFormulaGuardError):
        sort_excel_sheet(source, "Tutar", True, destination)

    assert not destination.exists()


def test_sort_excel_sheet_leaves_source_untouched_when_formula_guard_triggers(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "sorted.xlsx"
    _write_workbook_with_formula(
        source,
        header=["Ad", "Tutar"],
        data_rows=[["Ali", 100], ["Veli", 200]],
        formula_cell="C2",
        formula="=B2+B3",
    )
    original_bytes = source.read_bytes()

    with pytest.raises(ExcelSortFormulaGuardError):
        sort_excel_sheet(source, "Tutar", True, destination)

    assert source.read_bytes() == original_bytes


# --- sort_excel_sheet: real sort (no formulas) ---


def test_sort_excel_sheet_reorders_rows_ascending_by_target_column(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "sorted.xlsx"
    _write_workbook(
        source,
        [
            ["Ad", "Tutar"],
            ["Veli", 200],
            ["Ali", 100],
            ["Can", 300],
        ],
    )

    sort_excel_sheet(source, "Tutar", True, destination)

    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    values = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert values == [100, 200, 300]


def test_sort_excel_sheet_reorders_rows_descending_by_target_column(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "sorted.xlsx"
    _write_workbook(
        source,
        [
            ["Ad", "Tutar"],
            ["Veli", 200],
            ["Ali", 100],
            ["Can", 300],
        ],
    )

    sort_excel_sheet(source, "Tutar", False, destination)

    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    values = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert values == [300, 200, 100]


def test_sort_excel_sheet_does_not_modify_the_source_file(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "sorted.xlsx"
    _write_workbook(
        source,
        [
            ["Ad", "Tutar"],
            ["Veli", 200],
            ["Ali", 100],
        ],
    )
    original_bytes = source.read_bytes()

    sort_excel_sheet(source, "Tutar", True, destination)

    assert source.read_bytes() == original_bytes


def test_sort_excel_sheet_supports_bare_column_letter_when_no_header_matches(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "sorted.xlsx"
    _write_workbook(
        source,
        [
            ["Ad", "Tutar"],
            ["Veli", 200],
            ["Ali", 100],
        ],
    )

    sort_excel_sheet(source, "B", True, destination)

    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    values = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert values == [100, 200]


def test_sort_excel_sheet_raises_value_error_for_unknown_column(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "sorted.xlsx"
    _write_workbook(
        source,
        [
            ["Ad", "Tutar"],
            ["Veli", 200],
        ],
    )

    with pytest.raises(ValueError):
        sort_excel_sheet(source, "BilinmeyenSutun", True, destination)

    assert not destination.exists()


# --- sort_excel_sheet: empty data no-op (ATDD P2 #6) ---


def test_sort_excel_sheet_is_a_no_op_copy_when_only_header_row_exists(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "sorted.xlsx"
    _write_workbook(source, [["Ad", "Tutar"]])

    sort_excel_sheet(source, "Tutar", True, destination)

    assert destination.exists()
    wb = openpyxl.load_workbook(str(destination))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows == [("Ad", "Tutar")]


def test_sort_excel_sheet_is_a_no_op_copy_for_a_completely_empty_sheet(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "sorted.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(source))

    # Bos sayfada bile bir sutun spesifiği veriliyor olabilir; no-op
    # olmasi gereken durumda formul kontrolu/sutun cozumlemesi
    # ANLAMSIZ oldugu icin hata FIRLATILMAMALI (ATDD P2 #6).
    sort_excel_sheet(source, "A", True, destination)

    assert destination.exists()
