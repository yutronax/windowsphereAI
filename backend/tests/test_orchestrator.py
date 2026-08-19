import os
import shutil
import time
from pathlib import Path
from unittest.mock import patch

import pytest
from pydantic import ValidationError
from sqlalchemy.orm import Session

from backend.db import create_db_engine, create_session_factory
from backend.db_models import Base, Transaction
from backend.models import (
    DateSource,
    OperationType,
    PdfFileMetadata,
    PlanSkeleton,
    PlanStep,
    RedactionRegion,
    SortOrder,
)
import datetime as dt

from backend.orchestrator import (
    PlanApplicationError,
    TransactionRevertError,
    apply_plan,
    purge_expired_delete_backups,
    # purge_oversized_delete_backups,  # RED: function doesn't exist yet (will be imported after implementation)
    recover_incomplete_transactions,
    revert_transaction,
)
from backend.security import PathWhitelistError


@pytest.fixture
def session() -> Session:
    engine = create_db_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    factory = create_session_factory(engine)
    with factory() as db_session:
        yield db_session


def _plan(steps: list[PlanStep]) -> PlanSkeleton:
    return PlanSkeleton(steps=steps, dateSource=DateSource.CREATED_AT, sortOrder=SortOrder.ASCENDING)


def _step(
    order: int,
    target_folder: str,
    file_names: list[str],
    operation_type: OperationType = OperationType.MOVE,
    new_file_names: list[str] | None = None,
) -> PlanStep:
    return PlanStep(
        order=order,
        operationType=operation_type,
        targetFolder=target_folder,
        affectedFileCount=len(file_names),
        fileNames=file_names,
        newFileNames=new_file_names,
    )


def _write_pdf(root, filename: str) -> None:
    (root / filename).write_bytes(b"%PDF-1.4 fake")


def _write_real_pdf(root, filename: str, page_count: int) -> None:
    # Saga #304: MERGE testleri gercek pypdf ile olusturulmus gercek PDF'ler
    # gerektiriyor (sayfa sayisi dogrulamasi icin) - _write_pdf'in sahte
    # bayt icerigi pypdf.PdfReader ile acilamaz.
    from pypdf import PdfWriter

    writer = PdfWriter()
    for _ in range(page_count):
        writer.add_blank_page(width=200, height=200)
    with open(root / filename, "wb") as f:
        writer.write(f)
    writer.close()


def test_apply_plan_moves_files_into_target_folder_and_records_completed_operations(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "b.pdf"])])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "a.pdf").exists()
    assert not (tmp_path / "b.pdf").exists()
    assert (tmp_path / "2026-08" / "a.pdf").exists()
    assert (tmp_path / "2026-08" / "b.pdf").exists()
    assert transaction.status == "committed"
    assert len(transaction.operations) == 2
    assert all(op.status == "completed" for op in transaction.operations)


def test_apply_plan_splits_pdf_files_across_multiple_steps_by_file_names(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-09-01"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf"]), _step(1, "2026-09", ["b.pdf"])])

    apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "2026-08" / "a.pdf").exists()
    assert (tmp_path / "2026-09" / "b.pdf").exists()


def test_apply_plan_ignores_step_order_when_matching_files_by_name(session, tmp_path):
    # Saga #286: dağıtım artık pozisyonel DEĞİL, isimle eşleşiyor — steps'in
    # sırası veya pdf_files'ın sırası dosya-step eşleşmesini etkilemez.
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="b.pdf", createdAt="2026-09-01"),
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
    ]
    plan = _plan([_step(1, "2026-09", ["b.pdf"]), _step(0, "2026-08", ["a.pdf"])])

    apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "2026-08" / "a.pdf").exists()
    assert (tmp_path / "2026-09" / "b.pdf").exists()


def test_apply_plan_rejects_whole_plan_without_moving_anything_when_a_pdf_file_is_unassigned(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf"])])  # b.pdf hiçbir step'e atanmadı

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "a.pdf").exists()
    assert (tmp_path / "b.pdf").exists()
    assert not (tmp_path / "2026-08").exists()


def test_apply_plan_rejects_whole_plan_when_a_step_references_an_unknown_file(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "ghost.pdf"])])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "a.pdf").exists()
    assert not (tmp_path / "2026-08").exists()


def test_apply_plan_rejects_whole_plan_when_a_source_file_escapes_the_whitelist(session, tmp_path):
    pdf_files = [PdfFileMetadata(filename="..", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", [".."])])

    with pytest.raises(Exception):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "2026-08").exists()


def test_apply_plan_does_not_touch_any_file_for_a_list_only_plan(session, tmp_path):
    # Saga #291: LIST tamamen salt okunur/inert - hicbir dosya sistemi
    # cagrisi yapilmaz, hicbir FileOperation kaydi olusturulmaz. Bu test
    # onceden "desteklenmeyen operationType" ornegi olarak LIST kullanan
    # bir testin yerini aldi - artik OperationType enum'undaki TUM 5
    # deger destekleniyor, "desteklenmeyen tip" senaryosu artik yok.
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"], operation_type=OperationType.LIST)])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "a.pdf").exists()
    assert not (tmp_path / "2026-08").exists()
    assert transaction.status == "committed"
    assert transaction.operations == []


def test_apply_plan_only_processes_the_move_step_in_a_mixed_list_and_move_plan(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan(
        [
            _step(0, "2026-08", ["a.pdf"], operation_type=OperationType.LIST),
            _step(1, "2026-08", ["b.pdf"], operation_type=OperationType.MOVE),
        ]
    )

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "a.pdf").exists()  # LIST dokunmadi
    assert not (tmp_path / "b.pdf").exists()  # MOVE tasidi
    assert (tmp_path / "2026-08" / "b.pdf").exists()
    assert len(transaction.operations) == 1  # sadece MOVE icin kayit var


def test_apply_plan_rejects_an_operation_type_missing_from_supported_set(session, tmp_path, monkeypatch):
    # Red-team bulgusu (Saga #291): OperationType enum'undaki TUM 5
    # deger artik destekleniyor, bu yuzden _SUPPORTED_OPERATION_TYPES
    # kontrolu su an olu kod - gercek bir enum degeriyle tetiklenemiyor.
    # Gelecekte enum'a yeni bir deger eklenip _SUPPORTED_OPERATION_TYPES
    # guncellenmesi UNUTULURSA bu guard'in hala calistigini dogrulamak
    # icin _SUPPORTED_OPERATION_TYPES'i monkeypatch ile daraltiyoruz.
    import backend.orchestrator as orchestrator_module

    monkeypatch.setattr(orchestrator_module, "_SUPPORTED_OPERATION_TYPES", {OperationType.MOVE})
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"], operation_type=OperationType.COPY)])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "a.pdf").exists()
    assert not (tmp_path / "2026-08").exists()


def test_apply_plan_rolls_back_completed_moves_when_a_later_move_fails(session, tmp_path, monkeypatch):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "b.pdf"])])

    real_move = shutil.move
    call_count = {"n": 0}

    def flaky_move(src, dst):
        call_count["n"] += 1
        if call_count["n"] == 2:
            raise OSError("simulated disk failure")
        return real_move(src, dst)

    monkeypatch.setattr("backend.orchestrator.shutil.move", flaky_move)

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    # a.pdf taşındı, sonra rollback ile eski konumuna geri döndü.
    assert (tmp_path / "a.pdf").exists()
    assert not (tmp_path / "2026-08" / "a.pdf").exists()
    # b.pdf'in taşınması hiç başarılı olmadı, zaten yerinde kalmalı.
    assert (tmp_path / "b.pdf").exists()


def test_apply_plan_rolls_back_completed_moves_in_reverse_order(session, tmp_path, monkeypatch):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    _write_pdf(tmp_path, "c.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
        PdfFileMetadata(filename="c.pdf", createdAt="2026-08-03"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "b.pdf", "c.pdf"])])

    real_move = shutil.move
    move_order: list[str] = []

    def tracking_move(src, dst):
        move_order.append(f"{src}->{dst}")
        if len(move_order) == 3:  # c.pdf'in ileri taşınması başarısız olsun
            raise OSError("simulated disk failure")
        return real_move(src, dst)

    monkeypatch.setattr("backend.orchestrator.shutil.move", tracking_move)

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    # Rollback hareketleri, ileri taşımaların TERS sırasıyla gerçekleşmeli:
    # önce (son tamamlanan) b.pdf, sonra a.pdf geri taşınmalı.
    reverse_moves = move_order[3:]
    assert len(reverse_moves) == 2
    assert reverse_moves[0].split("->")[0].endswith("b.pdf")
    assert reverse_moves[1].split("->")[0].endswith("a.pdf")
    assert (tmp_path / "a.pdf").exists()
    assert (tmp_path / "b.pdf").exists()
    assert (tmp_path / "c.pdf").exists()
    assert not (tmp_path / "2026-08").exists() or list((tmp_path / "2026-08").iterdir()) == []


def test_apply_plan_reads_rollback_source_from_recorded_backup_path(session, tmp_path, monkeypatch):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "b.pdf"])])

    real_move = shutil.move
    call_count = {"n": 0}

    def flaky_move(src, dst):
        call_count["n"] += 1
        if call_count["n"] == 2:
            raise OSError("simulated disk failure")
        return real_move(src, dst)

    monkeypatch.setattr("backend.orchestrator.shutil.move", flaky_move)

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    completed_op = session.query(Transaction).one().operations[0]
    assert completed_op.backup_path == str(tmp_path / "a.pdf")
    assert completed_op.status == "rolled_back"
    assert (tmp_path / "a.pdf").exists()


def test_apply_plan_marks_transaction_and_operations_rolled_back_on_failure(session, tmp_path, monkeypatch):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "b.pdf"])])

    real_move = shutil.move
    call_count = {"n": 0}

    def flaky_move(src, dst):
        call_count["n"] += 1
        if call_count["n"] == 2:
            raise OSError("simulated disk failure")
        return real_move(src, dst)

    monkeypatch.setattr("backend.orchestrator.shutil.move", flaky_move)

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    transaction = session.query(Transaction).one()
    assert transaction.status == "rolled_back"
    assert all(op.status == "rolled_back" for op in transaction.operations)


def test_apply_plan_copies_files_leaving_the_source_intact(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"], operation_type=OperationType.COPY)])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "a.pdf").exists()  # kaynak hala yerinde
    assert (tmp_path / "2026-08" / "a.pdf").exists()  # kopya olusturuldu
    assert transaction.status == "committed"
    assert transaction.operations[0].status == "completed"


def test_apply_plan_rolls_back_a_copy_by_deleting_the_destination_without_touching_source(session, tmp_path, monkeypatch):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "b.pdf"], operation_type=OperationType.COPY)])

    real_copy = shutil.copy2
    call_count = {"n": 0}

    def flaky_copy(src, dst):
        call_count["n"] += 1
        if call_count["n"] == 2:
            raise OSError("simulated disk failure")
        return real_copy(src, dst)

    monkeypatch.setattr("backend.orchestrator.shutil.copy2", flaky_copy)

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    # a.pdf kopyalanmisti, rollback hedefteki kopyayi sildi.
    assert (tmp_path / "a.pdf").exists()  # kaynak hic dokunulmadi
    assert not (tmp_path / "2026-08" / "a.pdf").exists()  # kopya silindi
    assert (tmp_path / "b.pdf").exists()  # b.pdf hic kopyalanmadi


def test_apply_plan_raises_original_error_not_a_masked_one_when_rollback_dispatch_fails(session, tmp_path, monkeypatch):
    # Red-team bulgusu (Saga #288): rollback sozlugu aramasi (OperationType(...))
    # eskiden sadece OSError'a karsi korunuyordu - bilinmeyen bir operation_type
    # ValueError/KeyError firlatip orijinal hatayi (exc) maskeleyebilirdi ve
    # transaction sonsuza dek "pending" kalirdi. Bu artik yakalaniyor.
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "b.pdf"])])

    real_move = shutil.move
    call_count = {"n": 0}

    def flaky_move(src, dst):
        call_count["n"] += 1
        if call_count["n"] == 2:
            raise OSError("simulated disk failure")
        return real_move(src, dst)

    monkeypatch.setattr("backend.orchestrator.shutil.move", flaky_move)
    monkeypatch.setattr("backend.orchestrator._ROLLBACK_OPERATIONS", {})  # bilinmeyen operation_type simule eder

    with pytest.raises(PlanApplicationError, match="simulated disk failure"):
        apply_plan(session, plan, pdf_files, tmp_path)

    transaction = session.query(Transaction).one()
    assert transaction.status == "rolled_back"  # "pending" asili KALMADI
    assert transaction.operations[0].status == "rollback_failed"


def test_apply_plan_deletes_files_after_backing_them_up(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"], operation_type=OperationType.DELETE)])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "a.pdf").exists()  # kaynak silindi
    backup_path = tmp_path / ".windows-ai-files-backup" / str(transaction.id) / "a.pdf"
    assert backup_path.exists()  # gercek fiziksel yedek olusturuldu
    assert transaction.status == "committed"
    assert transaction.operations[0].status == "completed"


def test_apply_plan_rolls_back_a_delete_by_restoring_the_file_from_backup(session, tmp_path, monkeypatch):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "b.pdf"], operation_type=OperationType.DELETE)])

    real_unlink = Path.unlink
    call_count = {"n": 0}

    def flaky_unlink(self, *args, **kwargs):
        call_count["n"] += 1
        if call_count["n"] == 2:
            raise OSError("simulated disk failure")
        return real_unlink(self, *args, **kwargs)

    monkeypatch.setattr("pathlib.Path.unlink", flaky_unlink)

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    # a.pdf silinip yedeklendi, sonra rollback ile orijinal konumuna geri geldi.
    assert (tmp_path / "a.pdf").exists()
    # b.pdf silme islemi hic basarili olmadi, zaten yerinde kalmali.
    assert (tmp_path / "b.pdf").exists()


def test_delete_backup_path_stays_within_max_path_depth(tmp_path):
    from backend.orchestrator import _DELETE_BACKUP_DIRNAME, _delete_backup_path
    from backend.security import MAX_PATH_DEPTH, is_path_too_deep

    backup_path = _delete_backup_path(tmp_path, transaction_id=1, filename="a.pdf")

    assert backup_path.parent.parent.name == _DELETE_BACKUP_DIRNAME
    backup_path.parent.mkdir(parents=True)
    backup_path.touch()
    assert is_path_too_deep(backup_path, tmp_path, max_depth=MAX_PATH_DEPTH) is False


def test_apply_plan_renames_a_file_in_place(session, tmp_path):
    _write_pdf(tmp_path, "eski.pdf")
    pdf_files = [PdfFileMetadata(filename="eski.pdf", createdAt="2026-08-01")]
    plan = _plan(
        [_step(0, "2026-08", ["eski.pdf"], operation_type=OperationType.RENAME, new_file_names=["yeni.pdf"])]
    )

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "eski.pdf").exists()
    assert (tmp_path / "yeni.pdf").exists()
    assert transaction.status == "committed"


def test_apply_plan_rename_output_filename_changes_when_new_file_names_changes(session, tmp_path):
    _write_pdf(tmp_path, "one.pdf")
    _write_pdf(tmp_path, "two.pdf")
    _write_pdf(tmp_path, "three.pdf")
    pdf_files = [
        PdfFileMetadata(filename="one.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="two.pdf", createdAt="2026-08-02"),
        PdfFileMetadata(filename="three.pdf", createdAt="2026-08-03"),
    ]
    plan = _plan(
        [
            _step(0, "2026-08", ["one.pdf"], operation_type=OperationType.RENAME, new_file_names=["alpha.pdf"]),
            _step(1, "2026-08", ["two.pdf"], operation_type=OperationType.RENAME, new_file_names=["beta.pdf"]),
            _step(2, "2026-08", ["three.pdf"], operation_type=OperationType.RENAME, new_file_names=["gamma.pdf"]),
        ]
    )

    apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "alpha.pdf").exists()
    assert not (tmp_path / "one.pdf").exists()
    assert (tmp_path / "beta.pdf").exists()
    assert not (tmp_path / "two.pdf").exists()
    assert (tmp_path / "gamma.pdf").exists()
    assert not (tmp_path / "three.pdf").exists()


def test_apply_plan_rolls_back_a_rename_by_restoring_the_old_name(session, tmp_path, monkeypatch):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan(
        [
            _step(
                0,
                "2026-08",
                ["a.pdf", "b.pdf"],
                operation_type=OperationType.RENAME,
                new_file_names=["a2.pdf", "b2.pdf"],
            )
        ]
    )

    real_move = shutil.move
    call_count = {"n": 0}

    def flaky_move(src, dst):
        call_count["n"] += 1
        if call_count["n"] == 2:
            raise OSError("simulated disk failure")
        return real_move(src, dst)

    monkeypatch.setattr("backend.orchestrator.shutil.move", flaky_move)

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "a.pdf").exists()  # eski ismine geri geldi
    assert not (tmp_path / "a2.pdf").exists()
    assert (tmp_path / "b.pdf").exists()  # hic yeniden adlandirilmadi


def test_apply_plan_rejects_cross_step_chained_renames_without_losing_data(session, tmp_path):
    # 2. red-team turu bulgusu (deneysel dogrulandi, HIGH): apply_plan
    # seviyesinde ucdan uca dogrulama - a.pdf->b.pdf (step 0) VE
    # b.pdf->c.pdf (step 1) ayni planda olursa, b.pdf'in orijinal
    # icerigi (bu testte ayirt edici bir icerikle yazilmis) sessizce
    # kaybolmamali - plan hicbir dosyaya dokunmadan reddedilmeli.
    (tmp_path / "a.pdf").write_bytes(b"A-icerik")
    (tmp_path / "b.pdf").write_bytes(b"B-ORIJINAL-ONEMLI-VERI")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan(
        [
            _step(0, "2026-08", ["a.pdf"], operation_type=OperationType.RENAME, new_file_names=["b.pdf"]),
            _step(1, "2026-08", ["b.pdf"], operation_type=OperationType.RENAME, new_file_names=["c.pdf"]),
        ]
    )

    with pytest.raises(Exception):
        apply_plan(session, plan, pdf_files, tmp_path)

    # Hicbir dosyaya dokunulmamis olmali - b.pdf'in orijinal icerigi korunuyor.
    assert (tmp_path / "a.pdf").read_bytes() == b"A-icerik"
    assert (tmp_path / "b.pdf").read_bytes() == b"B-ORIJINAL-ONEMLI-VERI"
    assert not (tmp_path / "c.pdf").exists()


# Saga #304: MERGE operasyonu - N kaynak PDF -> 1 yeni birlesik PDF,
# kaynaklara dokunulmaz, rollback COPY semantigiyle (sadece hedefi sil).


def _merge_step(order: int, file_names: list[str], merged_file_name: str) -> PlanStep:
    return PlanStep(
        order=order,
        operationType=OperationType.MERGE,
        targetFolder="2026-08",
        affectedFileCount=len(file_names),
        fileNames=file_names,
        mergedFileName=merged_file_name,
    )


def test_apply_plan_merges_real_pdfs_into_one_file_with_the_correct_total_page_count(session, tmp_path):
    from pypdf import PdfReader

    _write_real_pdf(tmp_path, "a.pdf", 2)
    _write_real_pdf(tmp_path, "b.pdf", 3)
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_merge_step(0, ["a.pdf", "b.pdf"], "birlesik.pdf")])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    merged_path = tmp_path / "birlesik.pdf"
    assert merged_path.exists()
    reader = PdfReader(str(merged_path))
    assert len(reader.pages) == 5
    assert transaction.status == "committed"
    assert len(transaction.operations) == 1
    assert transaction.operations[0].status == "completed"


def test_apply_plan_merge_output_filename_changes_when_merged_file_name_changes(session, tmp_path):
    run1 = tmp_path / "run1"
    run1.mkdir()
    run2 = tmp_path / "run2"
    run2.mkdir()

    _write_real_pdf(run1, "s1.pdf", 1)
    _write_real_pdf(run1, "s2.pdf", 1)
    pdf_files_1 = [
        PdfFileMetadata(filename="s1.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="s2.pdf", createdAt="2026-08-02"),
    ]
    plan_1 = _plan([_merge_step(0, ["s1.pdf", "s2.pdf"], "merged_a.pdf")])
    apply_plan(session, plan_1, pdf_files_1, run1)

    assert (run1 / "merged_a.pdf").exists()

    _write_real_pdf(run2, "s3.pdf", 1)
    _write_real_pdf(run2, "s4.pdf", 1)
    pdf_files_2 = [
        PdfFileMetadata(filename="s3.pdf", createdAt="2026-08-03"),
        PdfFileMetadata(filename="s4.pdf", createdAt="2026-08-04"),
    ]
    plan_2 = _plan([_merge_step(0, ["s3.pdf", "s4.pdf"], "merged_b.pdf")])
    apply_plan(session, plan_2, pdf_files_2, run2)

    assert (run2 / "merged_b.pdf").exists()
    assert (run1 / "merged_a.pdf").exists()
    assert not (run2 / "merged_a.pdf").exists()
    assert not (run1 / "merged_b.pdf").exists()


def test_apply_plan_leaves_merge_sources_untouched(session, tmp_path):
    _write_real_pdf(tmp_path, "a.pdf", 1)
    _write_real_pdf(tmp_path, "b.pdf", 1)
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_merge_step(0, ["a.pdf", "b.pdf"], "birlesik.pdf")])

    apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "a.pdf").exists()
    assert (tmp_path / "b.pdf").exists()


def test_apply_plan_rolls_back_a_merge_by_deleting_the_merged_file_without_touching_sources(session, tmp_path):
    _write_real_pdf(tmp_path, "a.pdf", 1)
    _write_real_pdf(tmp_path, "b.pdf", 1)
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
        PdfFileMetadata(filename="c.pdf", createdAt="2026-08-03"),
    ]
    plan = _plan(
        [
            _merge_step(0, ["a.pdf", "b.pdf"], "birlesik.pdf"),
            _step(1, "2026-08", ["c.pdf"]),  # c.pdf pdf_files'ta ama diskte yok - bu step basarisiz olur
        ]
    )

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "birlesik.pdf").exists()
    assert (tmp_path / "a.pdf").exists()
    assert (tmp_path / "b.pdf").exists()


def test_apply_plan_merge_leaves_no_temp_files_behind_on_success(session, tmp_path):
    # Red-team bulgusu (Saga #304): _forward_merge artik gecici dosyaya
    # yazip ATOMIK rename ile hedefe tasiyor - basarili bir merge sonrasi
    # klasorde SADECE birlesik.pdf, a.pdf, b.pdf kalmali, hicbir ".tmp"
    # artik dosya birakilmamali.
    _write_real_pdf(tmp_path, "a.pdf", 1)
    _write_real_pdf(tmp_path, "b.pdf", 1)
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_merge_step(0, ["a.pdf", "b.pdf"], "birlesik.pdf")])

    apply_plan(session, plan, pdf_files, tmp_path)

    remaining = {p.name for p in tmp_path.iterdir()}
    assert remaining == {"a.pdf", "b.pdf", "birlesik.pdf"}


def test_forward_merge_leaves_no_partial_file_at_destination_when_write_fails(tmp_path, monkeypatch):
    # _forward_merge'i dogrudan cagirip PdfWriter.write'in yazma sirasinda
    # basarisiz oldugunu simule ediyoruz (disk dolu / izin hatasi benzeri).
    # Gecici-dosya + atomik-rename deseni sayesinde gercek destination_path'te
    # HICBIR yarim/bozuk dosya kalmamali.
    from pypdf import PdfWriter

    from backend.orchestrator import _forward_merge

    _write_real_pdf(tmp_path, "a.pdf", 1)
    _write_real_pdf(tmp_path, "b.pdf", 1)
    destination = tmp_path / "birlesik.pdf"

    def _boom(self, path):
        raise OSError("simulated disk full")

    monkeypatch.setattr(PdfWriter, "write", _boom)

    with pytest.raises(OSError):
        _forward_merge([tmp_path / "a.pdf", tmp_path / "b.pdf"], destination)

    assert not destination.exists()
    # Hicbir ".tmp" artik dosyasi da kalmamis olmali.
    leftovers = [p for p in tmp_path.iterdir() if p.name not in {"a.pdf", "b.pdf"}]
    assert leftovers == []


# Saga #305: SPLIT operasyonu - 1 kaynak PDF -> N yeni tek-sayfalik PDF,
# kaynaga dokunulmaz, rollback COPY semantigiyle (her cikti icin sadece
# hedefi sil).


def _split_step(order: int, file_name: str) -> PlanStep:
    return PlanStep(
        order=order,
        operationType=OperationType.SPLIT,
        targetFolder="2026-08",
        affectedFileCount=1,
        fileNames=[file_name],
    )


def test_apply_plan_splits_a_real_pdf_into_one_file_per_page(session, tmp_path):
    from pypdf import PdfReader

    _write_real_pdf(tmp_path, "rapor.pdf", 3)
    pdf_files = [PdfFileMetadata(filename="rapor.pdf", createdAt="2026-08-01")]
    plan = _plan([_split_step(0, "rapor.pdf")])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    for page_number in (1, 2, 3):
        output_path = tmp_path / f"rapor_{page_number}.pdf"
        assert output_path.exists()
        reader = PdfReader(str(output_path))
        assert len(reader.pages) == 1
    assert transaction.status == "committed"
    assert len(transaction.operations) == 3
    assert all(op.status == "completed" for op in transaction.operations)


def test_apply_plan_leaves_split_source_untouched(session, tmp_path):
    _write_real_pdf(tmp_path, "rapor.pdf", 2)
    pdf_files = [PdfFileMetadata(filename="rapor.pdf", createdAt="2026-08-01")]
    plan = _plan([_split_step(0, "rapor.pdf")])

    apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "rapor.pdf").exists()


def test_apply_plan_rejects_split_when_output_file_name_already_exists(session, tmp_path):
    _write_real_pdf(tmp_path, "rapor.pdf", 3)
    # rapor_2.pdf plan'in BILMEDIGI, onceden var olan ilgisiz bir dosya -
    # cikti adiyla CAKISIYOR.
    (tmp_path / "rapor_2.pdf").write_bytes(b"%PDF-1.4 onceden-var")
    pdf_files = [PdfFileMetadata(filename="rapor.pdf", createdAt="2026-08-01")]
    plan = _plan([_split_step(0, "rapor.pdf")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    # rapor_1.pdf (cakismadan ONCE yazilmis olabilecek) rollback ile
    # silinmis olmali - hicbir SPLIT ciktisi geride kalmamali.
    assert not (tmp_path / "rapor_1.pdf").exists()
    # Cakisan onceden-var-olan dosya DOKUNULMADAN kalmali (uzerine yazilmadi).
    assert (tmp_path / "rapor_2.pdf").read_bytes() == b"%PDF-1.4 onceden-var"
    assert not (tmp_path / "rapor_3.pdf").exists()
    assert (tmp_path / "rapor.pdf").exists()


def test_apply_plan_rejects_split_of_a_zero_page_pdf(session, tmp_path):
    _write_real_pdf(tmp_path, "bos.pdf", 0)
    pdf_files = [PdfFileMetadata(filename="bos.pdf", createdAt="2026-08-01")]
    plan = _plan([_split_step(0, "bos.pdf")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)


def test_apply_plan_rolls_back_split_outputs_when_a_later_step_fails(session, tmp_path):
    _write_real_pdf(tmp_path, "rapor.pdf", 2)
    pdf_files = [
        PdfFileMetadata(filename="rapor.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="c.pdf", createdAt="2026-08-03"),
    ]
    plan = _plan(
        [
            _split_step(0, "rapor.pdf"),
            _step(1, "2026-08", ["c.pdf"]),  # c.pdf pdf_files'ta ama diskte yok - basarisiz olur
        ]
    )

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "rapor_1.pdf").exists()
    assert not (tmp_path / "rapor_2.pdf").exists()
    assert (tmp_path / "rapor.pdf").exists()


# OCR operasyonu (red step, Saga: apply_plan henuz OperationType.OCR'i
# desteklemiyor, backend.orchestrator.ocr_pdf_file diye bir isim de
# import edilmis degil). Bu testler simdilik KIRMIZI kalmali.


def test_apply_plan_runs_ocr_on_a_pdf_inside_allowed_root_without_moving_it(session, tmp_path, monkeypatch):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"], operation_type=OperationType.OCR)])

    calls: list[Path] = []

    def fake_ocr_pdf_file(pdf_path: Path) -> list[str]:
        calls.append(pdf_path)
        return ["metin"]

    monkeypatch.setattr("backend.orchestrator.ocr_pdf_file", fake_ocr_pdf_file)

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    assert calls == [tmp_path / "a.pdf"]
    assert (tmp_path / "a.pdf").exists()
    assert transaction.status == "committed"
    assert len(transaction.operations) == 0


def test_apply_plan_rejects_ocr_of_a_path_outside_allowed_root(session, tmp_path, monkeypatch):
    # PdfFileMetadata.filename kendi validator'unda ayrac icermez, bu yuzden
    # dogrudan bir path-traversal stringi ifade edilemez - ama tek segmentlik
    # ".." (Saga #272'deki mevcut testlerin kullandigi teknikle AYNI) allowed_root
    # disina cikar. apply_plan'in OCR adiminin da diger operationType'lar
    # (MERGE/SPLIT/DELETE) gibi is_path_allowed/validate_plan_paths uzerinden
    # kaynagi dogrulamasi gerektigini, mock OCR fonksiyonunun HIC
    # cagrilmadigini dogrulayarak test ediyoruz.
    pdf_files = [PdfFileMetadata(filename="..", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", [".."], operation_type=OperationType.OCR)])

    calls: list[Path] = []

    def fake_ocr_pdf_file(pdf_path: Path) -> list[str]:
        calls.append(pdf_path)
        return ["metin"]

    monkeypatch.setattr("backend.orchestrator.ocr_pdf_file", fake_ocr_pdf_file)

    # Red-team notu (Saga #307): burada bilerek genel Exception yerine
    # PathWhitelistError bekleniyor - guvenlik testinin, gercek path
    # dogrulamasindan farkli bir hatayla (ör. bir yazim hatasindan gelen
    # AttributeError/KeyError) yanlislikla "yesil" gecmesini engellemek icin.
    with pytest.raises(PathWhitelistError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert calls == []


def test_recover_incomplete_transactions_marks_physically_moved_files_as_completed(session, tmp_path):
    # Saga #286: shutil.move basarili oldu ama surec operation.status =
    # "completed" + commit'ten ONCE coktu senaryosunu simule eder - kayit
    # "pending" kalir, dosya fiziksel olarak hedefte durur.
    target_dir = tmp_path / "2026-08"
    target_dir.mkdir()
    (target_dir / "a.pdf").write_bytes(b"%PDF-1.4 fake")

    from backend.file_operations import create_transaction, record_file_operation

    transaction = create_transaction(session)
    record_file_operation(
        session,
        transaction,
        operation_type="Taşı",
        source_path=str(tmp_path / "a.pdf"),
        destination_path=str(target_dir / "a.pdf"),
        backup_path=str(tmp_path / "a.pdf"),
    )
    transaction.status = "pending"
    session.commit()

    recovered = recover_incomplete_transactions(session)

    assert len(recovered) == 1
    assert recovered[0].status == "committed"
    assert recovered[0].operations[0].status == "completed"


def test_recover_incomplete_transactions_marks_never_moved_files_as_rolled_back(session, tmp_path):
    from backend.file_operations import create_transaction, record_file_operation

    transaction = create_transaction(session)
    record_file_operation(
        session,
        transaction,
        operation_type="Taşı",
        source_path=str(tmp_path / "a.pdf"),
        destination_path=str(tmp_path / "2026-08" / "a.pdf"),  # hiç oluşmadı
        backup_path=str(tmp_path / "a.pdf"),
    )
    transaction.status = "pending"
    session.commit()

    recovered = recover_incomplete_transactions(session)

    assert len(recovered) == 1
    assert recovered[0].status == "rolled_back"
    assert recovered[0].operations[0].status == "rolled_back"


def test_recover_incomplete_transactions_reverifies_stale_completed_operations_in_a_pending_transaction(session, tmp_path):
    # Saga #286 red-team bulgusu: apply_plan'in rollback except bloğu bir
    # operasyonu bellek-içi "completed"->"rolled_back"e cevirdikten SONRA
    # ama nihai session.commit()'ten ONCE surec cokerse, DB'de o operasyon
    # hala "completed" gorunur - ama dosya fiziksel olarak zaten geri
    # tasinmis olabilir. Sadece status=="pending" olanlari kontrol etmek
    # bu operasyonu sonsuza dek yanlis etiketli birakirdi.
    from backend.file_operations import create_transaction, record_file_operation

    transaction = create_transaction(session)
    stale_completed_op = record_file_operation(
        session,
        transaction,
        operation_type="Taşı",
        source_path=str(tmp_path / "a.pdf"),
        destination_path=str(tmp_path / "2026-08" / "a.pdf"),  # fiziksel olarak geri tasindi, artik yok
        backup_path=str(tmp_path / "a.pdf"),
    )
    stale_completed_op.status = "completed"  # DB'de hala "completed", ama dosya orada degil
    still_pending_op = record_file_operation(
        session,
        transaction,
        operation_type="Taşı",
        source_path=str(tmp_path / "b.pdf"),
        destination_path=str(tmp_path / "2026-08" / "b.pdf"),  # hic tasinmadi
        backup_path=str(tmp_path / "b.pdf"),
    )
    transaction.status = "pending"
    session.commit()

    recovered = recover_incomplete_transactions(session)

    assert len(recovered) == 1
    refreshed_stale_op = next(op for op in recovered[0].operations if op.id == stale_completed_op.id)
    refreshed_pending_op = next(op for op in recovered[0].operations if op.id == still_pending_op.id)
    assert refreshed_stale_op.status == "rolled_back"
    assert refreshed_pending_op.status == "rolled_back"
    assert recovered[0].status == "rolled_back"


def test_recover_incomplete_transactions_ignores_already_terminal_transactions(session, tmp_path):
    from backend.file_operations import create_transaction, record_file_operation

    transaction = create_transaction(session)
    record_file_operation(
        session,
        transaction,
        operation_type="Taşı",
        source_path=str(tmp_path / "a.pdf"),
        destination_path=str(tmp_path / "2026-08" / "a.pdf"),
        backup_path=str(tmp_path / "a.pdf"),
    )
    transaction.status = "committed"
    session.commit()

    recovered = recover_incomplete_transactions(session)

    assert recovered == []


def test_revert_transaction_moves_a_completed_move_back_to_its_original_location(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"])])
    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    reverted = revert_transaction(session, transaction)

    assert (tmp_path / "a.pdf").exists()
    assert not (tmp_path / "2026-08" / "a.pdf").exists()
    assert reverted.status == "reverted"
    assert all(op.status == "rolled_back" for op in reverted.operations)


def test_revert_transaction_deletes_the_copy_and_leaves_the_original_intact(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"], operation_type=OperationType.COPY)])
    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    revert_transaction(session, transaction)

    assert (tmp_path / "a.pdf").exists()
    assert not (tmp_path / "2026-08" / "a.pdf").exists()


def test_revert_transaction_restores_a_deleted_file_from_its_backup(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"], operation_type=OperationType.DELETE)])
    transaction = apply_plan(session, plan, pdf_files, tmp_path)
    assert not (tmp_path / "a.pdf").exists()

    revert_transaction(session, transaction)

    assert (tmp_path / "a.pdf").exists()


def test_revert_transaction_restores_the_old_name_of_a_renamed_file(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"], operation_type=OperationType.RENAME, new_file_names=["b.pdf"])])
    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    revert_transaction(session, transaction)

    assert (tmp_path / "a.pdf").exists()
    assert not (tmp_path / "b.pdf").exists()


def test_revert_transaction_reverts_multiple_operations_in_reverse_order(session, tmp_path, monkeypatch):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "b.pdf"])])
    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    revert_order: list[str] = []
    import backend.orchestrator as orchestrator_module

    original = orchestrator_module._rollback_move

    def _tracking_rollback_move(destination_path, backup_path):
        revert_order.append(Path(destination_path).name)
        original(destination_path, backup_path)

    monkeypatch.setattr(orchestrator_module, "_rollback_move", _tracking_rollback_move)
    monkeypatch.setitem(orchestrator_module._ROLLBACK_OPERATIONS, OperationType.MOVE, _tracking_rollback_move)

    revert_transaction(session, transaction)

    assert revert_order == ["b.pdf", "a.pdf"]


def test_revert_transaction_rejects_a_transaction_that_is_not_committed(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"])])
    transaction = apply_plan(session, plan, pdf_files, tmp_path)
    transaction.status = "reverted"
    session.commit()

    with pytest.raises(TransactionRevertError):
        revert_transaction(session, transaction)

    # Reddedilince hiçbir dosyaya dokunulmaz — dosya hâlâ apply_plan'ın
    # bıraktığı hedef klasörde, kök klasöre GERİ TAŞINMAMIŞ olmalı.
    assert not (tmp_path / "a.pdf").exists()
    assert (tmp_path / "2026-08" / "a.pdf").exists()


def test_revert_transaction_marks_revert_failed_and_raises_when_a_step_cannot_be_reverted(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan([_step(0, "2026-08", ["a.pdf", "b.pdf"])])
    transaction = apply_plan(session, plan, pdf_files, tmp_path)
    # b.pdf'in hedefteki kopyasını rollback'ten ÖNCE kaldırıp yerine bir
    # KLASÖR koyuyoruz — shutil.move hedef zaten var olan bir klasörse
    # OSError fırlatır, bu da rollback'in KENDİSİNİN başarısız olduğu
    # senaryoyu gerçekçi şekilde tetikler.
    b_destination = tmp_path / "2026-08" / "b.pdf"
    b_destination.unlink()
    b_destination.mkdir()
    # b.pdf'in eski konumunda da bir dosya bırakıyoruz ki shutil.move hedefi
    # zaten dolu bir konuma taşımaya çalışsın (Windows'ta bu OSError verir).
    _write_pdf(tmp_path, "b.pdf")

    with pytest.raises(TransactionRevertError):
        revert_transaction(session, transaction)

    assert transaction.status == "revert_failed"
    # a.pdf başarıyla geri alındı (b.pdf'ten SONRA işlendiği için ters
    # sırada a.pdf İLK geri alınır) — kısmi ilerleme kaybolmadı.
    assert (tmp_path / "a.pdf").exists()


def test_revert_transaction_ignores_operations_outside_the_allowed_root(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"])])
    transaction = apply_plan(session, plan, pdf_files, tmp_path)
    # DB'deki kaydı, allowed_root DIŞINDA bir yeri işaret edecek şekilde
    # elle bozuyoruz (Saga #293 S4: savunma derinliği senaryosu).
    outside_root = tmp_path.parent / "baska-bir-klasor"
    transaction.operations[0].destination_path = str(outside_root / "a.pdf")
    session.commit()

    with pytest.raises(TransactionRevertError):
        revert_transaction(session, transaction)

    assert transaction.status == "revert_failed"
    assert not outside_root.exists()


def _apply_a_delete_plan(session, tmp_path, filename: str = "a.pdf"):
    _write_pdf(tmp_path, filename)
    pdf_files = [PdfFileMetadata(filename=filename, createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", [filename], operation_type=OperationType.DELETE)])
    return apply_plan(session, plan, pdf_files, tmp_path)


def test_purge_expired_delete_backups_deletes_the_backup_folder_and_marks_the_transaction_purged(session, tmp_path):
    transaction = _apply_a_delete_plan(session, tmp_path)
    backup_dir = tmp_path / ".windows-ai-files-backup" / str(transaction.id)
    assert backup_dir.exists()
    transaction.created_at = dt.datetime.utcnow() - dt.timedelta(days=31)
    session.commit()

    purged_ids = purge_expired_delete_backups(session, tmp_path, older_than_days=30)

    assert purged_ids == [transaction.id]
    assert not backup_dir.exists()
    assert transaction.status == "backup_purged"


def test_purge_expired_delete_backups_ignores_transactions_newer_than_the_cutoff(session, tmp_path):
    transaction = _apply_a_delete_plan(session, tmp_path)
    backup_dir = tmp_path / ".windows-ai-files-backup" / str(transaction.id)

    purged_ids = purge_expired_delete_backups(session, tmp_path, older_than_days=30)

    assert purged_ids == []
    assert backup_dir.exists()
    assert transaction.status == "committed"


def test_purge_expired_delete_backups_skips_mixed_operation_transactions(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    _write_pdf(tmp_path, "b.pdf")
    pdf_files = [
        PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="b.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan(
        [
            _step(0, "2026-08", ["a.pdf"], operation_type=OperationType.DELETE),
            _step(1, "2026-08", ["b.pdf"], operation_type=OperationType.MOVE),
        ]
    )
    transaction = apply_plan(session, plan, pdf_files, tmp_path)
    backup_dir = tmp_path / ".windows-ai-files-backup" / str(transaction.id)
    assert backup_dir.exists()
    transaction.created_at = dt.datetime.utcnow() - dt.timedelta(days=31)
    session.commit()

    purged_ids = purge_expired_delete_backups(session, tmp_path, older_than_days=30)

    assert purged_ids == []
    assert backup_dir.exists()
    assert transaction.status == "committed"


def test_purge_expired_delete_backups_does_not_touch_a_transaction_belonging_to_a_different_root(
    session, tmp_path
):
    # Red-team bulgusu (HIGH): `Transaction` hangi `allowed_root`a ait
    # olduğunu bilmiyor — DB sorgusu TÜM köklerdeki committed transaction'ları
    # döndürür. Başka bir köke ait bir transaction için bu fonksiyona YANLIŞ
    # bir `allowed_root` verilirse (çok-kök senaryosu), o transaction'a
    # DOKUNULMAMALI (durumu `"committed"` kalmalı, backup'ı olduğu gibi
    # durmalı) — aksi halde hiçbir şey fiziksel olarak silinmeden geri
    # alınamaz hale gelirdi.
    other_root = tmp_path / "baska-bir-kok"
    other_root.mkdir()
    transaction = _apply_a_delete_plan(session, other_root)
    transaction.created_at = dt.datetime.utcnow() - dt.timedelta(days=31)
    session.commit()

    wrong_root = tmp_path / "yanlis-kok"
    wrong_root.mkdir()
    purged_ids = purge_expired_delete_backups(session, wrong_root, older_than_days=30)

    assert purged_ids == []
    assert transaction.status == "committed"
    assert (other_root / ".windows-ai-files-backup" / str(transaction.id)).exists()


def test_purge_expired_delete_backups_result_is_rejected_by_revert_transaction_without_any_code_change(
    session, tmp_path
):
    # Saga #300'ün asıl güvenlik gerekçesi: purge edilmiş bir transaction
    # `revert_transaction`e verilirse, `_rollback_completed_operations`in
    # "hedef fiziksel olarak yoksa zaten geri alınmış" kısayolu SESSİZCE
    # "başarılı" görünürdü — `"backup_purged"` durumu bunu `revert_transaction`in
    # ZATEN VAR OLAN committed-only guard'ıyla (Saga #293) engelliyor.
    transaction = _apply_a_delete_plan(session, tmp_path)
    transaction.created_at = dt.datetime.utcnow() - dt.timedelta(days=31)
    session.commit()
    purge_expired_delete_backups(session, tmp_path, older_than_days=30)

    with pytest.raises(TransactionRevertError):
        revert_transaction(session, transaction)

    # Reddedilince hiçbir "başarılı geri alma" görünümü yok, dosya
    # gerçekten kaybolmuş durumda kalıyor (bu testin amacı budur).
    assert not (tmp_path / "a.pdf").exists()


# Saga #301 (RED step): Transaction.allowed_root kolonu + revert_transaction'ın
# istemciden allowed_root almayıp kendi kayıtlı değerini kullanması.


def test_apply_plan_records_the_real_allowed_root_on_the_transaction(session, tmp_path):
    _write_pdf(tmp_path, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"])])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    assert transaction.allowed_root == str(tmp_path)


def test_revert_transaction_raises_when_allowed_root_is_none(session, tmp_path):
    from backend.file_operations import create_transaction

    transaction = create_transaction(session, allowed_root=None)
    transaction.status = "committed"
    session.commit()

    with pytest.raises(TransactionRevertError):
        revert_transaction(session, transaction)


def test_revert_transaction_signature_no_longer_accepts_a_client_supplied_allowed_root(session, tmp_path):
    # Guvenlik regresyon testi (Saga #301): revert_transaction artik SADECE
    # (session, transaction) aliyor - istemcinin spoof edebilecegi ayri bir
    # allowed_root parametresi YOK. Gercek kok, transaction'in KENDI
    # apply_plan sirasinda kaydedilen degeridir, cagiran tarafindan
    # degistirilemez.
    real_root = tmp_path
    _write_pdf(real_root, "a.pdf")
    pdf_files = [PdfFileMetadata(filename="a.pdf", createdAt="2026-08-01")]
    plan = _plan([_step(0, "2026-08", ["a.pdf"])])
    transaction = apply_plan(session, plan, pdf_files, real_root)

    assert transaction.allowed_root == str(real_root)

    reverted = revert_transaction(session, transaction)

    assert reverted.status == "reverted"
    assert (real_root / "a.pdf").exists()


# Saga #302 (RED step): revert_transaction ve purge_expired_delete_backups
# arasindaki lost-update yarisi. Gercek OS-thread'ler yerine, AYNI sqlite
# dosyasina baglanan iki BAGIMSIZ Session nesnesi kullanip, hangi
# session'in "once" commit ettigini elle interleave ederek simule ediyoruz.
# `session` fixture'i ":memory:" kullaniyor (her baglanti kendi ayri
# bellek-ici DB'sini gorur), bu yuzden bu test kendi dosya-tabanli engine'ini
# kuruyor - aynen fixture'in yaptigi create_db_engine/create_session_factory
# cagrilarini, ama gercek bir dosya yolu ile, boylece iki session AYNI
# satirlari gorebilsin.


def test_revert_transaction_and_purge_expired_delete_backups_do_not_lose_the_race(tmp_path, monkeypatch):
    from sqlalchemy import update

    import backend.orchestrator as orchestrator_module

    db_path = tmp_path / "race.db"
    engine = create_db_engine(f"sqlite:///{db_path}")
    factory = create_session_factory(engine)
    root = tmp_path / "root"
    root.mkdir()

    with factory() as session_a, factory() as session_b:
        # --- Senaryo 1: "Revert, satir az once purge edilmis" -------------
        # session_a DELETE-only bir transaction'i revert etmeye baslar;
        # revert'in dosya-geri-alma isini bitirip KENDI son commit'ini
        # yapmasindan HEMEN ONCE, session_b (purge_expired_delete_backups'in
        # claim'ini simule ederek) satiri "committed" -> "backup_purged"
        # yapip commit eder. Hedeflenen davranis: revert bunu fark edip
        # TransactionRevertError firlatmali ve KENDI "reverted" durumunu
        # DB'ye YAZMAMALI (satir "backup_purged" olarak kalmali). Mevcut
        # (henuz duzeltilmemis) kod bunu kontrol etmiyor, revert dosyalari
        # zaten geri almis olacak VE session_a'nin son commit'i session_b'nin
        # "backup_purged" yazisinin UZERINE sessizce "reverted" yazacak -
        # bu da testin RED olarak basarisiz olmasini saglar.
        transaction = _apply_a_delete_plan(session_a, root, "a.pdf")
        txn_id = transaction.id
        backup_dir = root / ".windows-ai-files-backup" / str(txn_id)
        assert backup_dir.exists()

        original_rollback = orchestrator_module._rollback_completed_operations

        def _rollback_then_let_purge_win_the_race(*args, **kwargs):
            result = original_rollback(*args, **kwargs)
            session_b.execute(
                update(Transaction)
                .where(Transaction.id == txn_id, Transaction.status == "committed")
                .values(status="backup_purged")
            )
            session_b.commit()
            return result

        monkeypatch.setattr(
            orchestrator_module, "_rollback_completed_operations", _rollback_then_let_purge_win_the_race
        )

        with pytest.raises(TransactionRevertError):
            revert_transaction(session_a, transaction)

        monkeypatch.undo()

        session_b.expire_all()
        final = session_b.get(Transaction, txn_id)
        # Kaybeden (revert) satiri KENDI sonucuyla EZMEMELI - kazananin
        # ("backup_purged") yazdigi deger korunmali.
        assert final.status == "backup_purged"

        # --- Senaryo 2: "Purge, satir az once revert edilmeye baslanmis" --
        # session_a onceden BASKA bir DELETE-only transaction icin revert
        # claim'ini kazanmis GIBI davranir ("committed" -> "reverting");
        # purge_expired_delete_backups (session_b) bu satiri aday olarak
        # secip KENDI atomik claim'ini ("committed" -> "purging") denemeden
        # HEMEN ONCE, session_a'nin claim'i devreye girer. Purge'un claim'i
        # artik `_claim_transaction_status` uzerinden ILK adimda (rmtree'den
        # ONCE calisir, Saga #302 red-team duzeltmesi: kilit artik rmtree
        # boyunca acik tutulmuyor) gerceklestigi icin, race'i bu fonksiyonun
        # ILK cagrisinin etrafinda simule ediyoruz. Hedeflenen davranis:
        # purge bu adayi ATLAMALI - purged_ids'e girmemeli, backup klasoru
        # SILINMEMELI.
        transaction2 = _apply_a_delete_plan(session_a, root, "b.pdf")
        txn2_id = transaction2.id
        backup_dir2 = root / ".windows-ai-files-backup" / str(txn2_id)
        assert backup_dir2.exists()
        transaction2.created_at = dt.datetime.utcnow() - dt.timedelta(days=31)
        session_a.commit()

        original_claim = orchestrator_module._claim_transaction_status
        claim_call_count = {"n": 0}

        def _claim_via_revert_then_real_claim(session, transaction_id, *, from_status, to_status):
            claim_call_count["n"] += 1
            if claim_call_count["n"] == 1 and transaction_id == txn2_id:
                session_a.execute(
                    update(Transaction)
                    .where(Transaction.id == txn2_id, Transaction.status == "committed")
                    .values(status="reverting")
                )
                session_a.commit()
            return original_claim(session, transaction_id, from_status=from_status, to_status=to_status)

        monkeypatch.setattr(orchestrator_module, "_claim_transaction_status", _claim_via_revert_then_real_claim)

        purged_ids = purge_expired_delete_backups(session_b, root, older_than_days=30)

        monkeypatch.undo()

        assert txn2_id not in purged_ids
        assert backup_dir2.exists()


# Saga #320: REDACT operasyonu - MERGE/SPLIT ile AYNI "tek kaynak PDF, COPY
# semantigiyle rollback (sadece hedefi sil)" deseni, ama N kaynak -> 1 hedef
# (MERGE) ya da 1 kaynak -> N hedef (SPLIT) DEGIL - TAM OLARAK 1 kaynak -> 1
# hedef, hedef sayfanin rasterize edilip uzerine opak siyah kutu(lar)
# cizilmis hali disindaki TUM sayfalar vektorel olarak KORUNUR. Bu testler
# simdilik KIRMIZI kalmali - `OperationType.REDACT`, `RedactionRegion`,
# `PlanStep.redactionRegions`/`redactedFileName`, `backend.pdf_redact` ve
# orchestrator'daki REDACT dali henuz YOK.
#
# `reportlab` bu venv'de kurulu degil - gercek cikarilabilir metin iceren
# PDF'ler, `test_pdf_redact.py`deki `_build_pdf_bytes` ile AYNI teknikle
# (standart PDF nesne modelini dogrudan bayt olarak el ile insa ederek)
# uretiliyor (deneysel dogrulandi: `pypdf.PdfReader(...).extract_text()`
# bu PDF'lerden metni basariyla cikarabiliyor).


def _build_text_pdf_bytes(page_texts: list[str]) -> bytes:
    n = len(page_texts)
    font_obj = 3
    first_page_obj = 4
    first_content_obj = 4 + n

    objects: dict[int, bytes] = {}
    objects[1] = b"<< /Type /Catalog /Pages 2 0 R >>"
    kids = " ".join(f"{first_page_obj + i} 0 R" for i in range(n))
    objects[2] = f"<< /Type /Pages /Kids [{kids}] /Count {n} >>".encode()
    objects[3] = b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"

    for i in range(n):
        page_num = first_page_obj + i
        content_num = first_content_obj + i
        objects[page_num] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 300 300] "
            f"/Resources << /Font << /F1 {font_obj} 0 R >> >> "
            f"/Contents {content_num} 0 R >>"
        ).encode()

    for i, text in enumerate(page_texts):
        content_num = first_content_obj + i
        escaped = text.replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")
        stream = f"BT /F1 24 Tf 20 150 Td ({escaped}) Tj ET".encode()
        objects[content_num] = (
            f"<< /Length {len(stream)} >>\nstream\n".encode() + stream + b"\nendstream"
        )

    out = bytearray()
    out += b"%PDF-1.4\n"
    offsets: dict[int, int] = {}
    for obj_num in sorted(objects):
        offsets[obj_num] = len(out)
        out += f"{obj_num} 0 obj\n".encode()
        out += objects[obj_num]
        out += b"\nendobj\n"

    xref_offset = len(out)
    max_obj = max(objects)
    out += f"xref\n0 {max_obj + 1}\n".encode()
    out += b"0000000000 65535 f \n"
    for obj_num in range(1, max_obj + 1):
        out += f"{offsets.get(obj_num, 0):010d} 00000 n \n".encode()
    out += b"trailer\n"
    out += f"<< /Size {max_obj + 1} /Root 1 0 R >>\n".encode()
    out += b"startxref\n"
    out += f"{xref_offset}\n".encode()
    out += b"%%EOF"
    return bytes(out)


def _write_text_pdf(root, filename: str, page_texts: list[str]) -> None:
    (root / filename).write_bytes(_build_text_pdf_bytes(page_texts))


def _redact_step(
    order: int,
    file_name: str,
    redacted_file_name: str,
    regions: list[RedactionRegion] | None = None,
) -> PlanStep:
    return PlanStep(
        order=order,
        operationType=OperationType.REDACT,
        targetFolder="2026-08",
        affectedFileCount=1,
        fileNames=[file_name],
        redactionRegions=regions or [RedactionRegion(page=1, x0=0, y0=0, x1=300, y1=300)],
        redactedFileName=redacted_file_name,
    )


REDACT_SENSITIVE_TEXT = "12345678901"


def test_apply_plan_redacts_a_region_and_the_output_no_longer_contains_the_original_text(session, tmp_path):
    from pypdf import PdfReader

    _write_text_pdf(
        tmp_path,
        "gizli.pdf",
        [
            f"TC Kimlik No: {REDACT_SENSITIVE_TEXT}",
            "ikinci sayfa - bu metin karartilmiyor",
        ],
    )
    source_size_before = (tmp_path / "gizli.pdf").stat().st_size
    pdf_files = [PdfFileMetadata(filename="gizli.pdf", createdAt="2026-08-01")]
    plan = _plan(
        [_redact_step(0, "gizli.pdf", "gizli_karartilmis.pdf", [RedactionRegion(page=1, x0=0, y0=0, x1=300, y1=300)])]
    )

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    output_path = tmp_path / "gizli_karartilmis.pdf"
    assert output_path.exists()
    reader = PdfReader(str(output_path))
    assert len(reader.pages) == 2
    assert REDACT_SENSITIVE_TEXT not in reader.pages[0].extract_text()
    # Karartilmayan diger sayfa hala metin-aranabilir olmali (rasterize
    # EDILMEMIS, vektorel olarak korunmus).
    assert "ikinci sayfa" in reader.pages[1].extract_text()
    # Rasterize edilmis (goruntu iceren) bir sayfa, kucuk bir metin-only
    # sayfadan cok daha buyuk cikti uretir.
    assert output_path.stat().st_size > source_size_before
    assert transaction.status == "committed"
    assert len(transaction.operations) == 1
    assert transaction.operations[0].status == "completed"


def test_apply_plan_leaves_redact_source_untouched(session, tmp_path):
    _write_text_pdf(tmp_path, "gizli.pdf", [f"TC: {REDACT_SENSITIVE_TEXT}"])
    original_bytes = (tmp_path / "gizli.pdf").read_bytes()
    pdf_files = [PdfFileMetadata(filename="gizli.pdf", createdAt="2026-08-01")]
    plan = _plan([_redact_step(0, "gizli.pdf", "gizli_karartilmis.pdf")])

    apply_plan(session, plan, pdf_files, tmp_path)

    assert (tmp_path / "gizli.pdf").exists()
    assert (tmp_path / "gizli.pdf").read_bytes() == original_bytes


def test_apply_plan_rejects_redact_of_a_path_outside_allowed_root(session, tmp_path, monkeypatch):
    # OCR testindeki AYNI ".." tekniği (Saga #307) - PdfFileMetadata.filename
    # kendi validator'unda ayrac icermez ama tek segmentlik ".." allowed_root
    # disina cikar. REDACT'in de diger operationType'lar gibi
    # is_path_allowed/validate_plan_paths uzerinden dogrulanmasi gerekir.
    pdf_files = [PdfFileMetadata(filename="..", createdAt="2026-08-01")]
    plan = _plan([_redact_step(0, "..", "gizli_karartilmis.pdf")])

    calls: list[Path] = []

    def fake_redact_pdf_page(pdf_path, page_number, regions):
        calls.append(pdf_path)
        return b"%PDF-1.4 fake-redacted"

    monkeypatch.setattr("backend.orchestrator.redact_pdf_page", fake_redact_pdf_page)

    with pytest.raises(PathWhitelistError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert calls == []
    assert not (tmp_path / "gizli_karartilmis.pdf").exists()


def test_apply_plan_rejects_redact_when_page_number_exceeds_page_count(session, tmp_path):
    _write_real_pdf(tmp_path, "kisa.pdf", 1)
    pdf_files = [PdfFileMetadata(filename="kisa.pdf", createdAt="2026-08-01")]
    plan = _plan(
        [_redact_step(0, "kisa.pdf", "kisa_karartilmis.pdf", [RedactionRegion(page=5, x0=0, y0=0, x1=100, y1=100)])]
    )

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "kisa_karartilmis.pdf").exists()
    assert (tmp_path / "kisa.pdf").exists()


def test_apply_plan_rolls_back_a_redact_by_deleting_the_output_file_without_touching_the_source(session, tmp_path):
    _write_text_pdf(tmp_path, "gizli.pdf", [f"TC: {REDACT_SENSITIVE_TEXT}"])
    original_bytes = (tmp_path / "gizli.pdf").read_bytes()
    pdf_files = [
        PdfFileMetadata(filename="gizli.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="c.pdf", createdAt="2026-08-03"),
    ]
    plan = _plan(
        [
            _redact_step(0, "gizli.pdf", "gizli_karartilmis.pdf"),
            _step(1, "2026-08", ["c.pdf"]),  # c.pdf pdf_files'ta ama diskte yok - bu step basarisiz olur
        ]
    )

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "gizli_karartilmis.pdf").exists()
    assert (tmp_path / "gizli.pdf").exists()
    assert (tmp_path / "gizli.pdf").read_bytes() == original_bytes


# Saga #324: EXCEL_SORT operasyonu (red step) - `OperationType.EXCEL_SORT`,
# `PlanStep.sortColumn`/`sortAscending`/`sortedFileName`, `backend.excel_sort`
# ve orchestrator'daki EXCEL_SORT dali henuz YOK. Bu testler simdilik
# KIRMIZI kalmali (AttributeError / ValidationError / ModuleNotFoundError -
# hepsi beklenen red durumu).


def _write_excel(root, filename: str, rows: list[list]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(str(root / filename))


def _write_excel_with_formula(
    root, filename: str, header: list, data_rows: list[list], formula_cell: str, formula: str
) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in data_rows:
        ws.append(row)
    ws[formula_cell] = formula
    wb.save(str(root / filename))


def _excel_sort_step(
    order: int,
    file_name: str,
    sort_column: str,
    sort_ascending: bool,
    sorted_file_name: str,
) -> PlanStep:
    return PlanStep(
        order=order,
        operationType=OperationType.EXCEL_SORT,
        targetFolder="2026-08",
        affectedFileCount=1,
        fileNames=[file_name],
        sortColumn=sort_column,
        sortAscending=sort_ascending,
        sortedFileName=sorted_file_name,
    )


def test_apply_plan_rejects_excel_sort_when_a_data_row_contains_a_formula(session, tmp_path):
    _write_excel_with_formula(
        tmp_path,
        "kaynak.xlsx",
        header=["Ad", "Tutar"],
        data_rows=[["Ali", 100], ["Veli", 200]],
        formula_cell="C2",
        formula="=B2+B3",
    )
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_sort_step(0, "kaynak.xlsx", "Tutar", True, "siralanmis.xlsx")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "siralanmis.xlsx").exists()


def test_apply_plan_sorts_excel_rows_when_there_are_no_formulas(session, tmp_path):
    import openpyxl

    _write_excel(
        tmp_path,
        "kaynak.xlsx",
        [
            ["Ad", "Tutar"],
            ["Veli", 200],
            ["Ali", 100],
            ["Can", 300],
        ],
    )
    original_bytes = (tmp_path / "kaynak.xlsx").read_bytes()
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_sort_step(0, "kaynak.xlsx", "Tutar", True, "siralanmis.xlsx")])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    output_path = tmp_path / "siralanmis.xlsx"
    assert output_path.exists()
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb.active
    values = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert values == [100, 200, 300]
    # Kaynak dosya asla degismemeli.
    assert (tmp_path / "kaynak.xlsx").read_bytes() == original_bytes
    assert transaction.status == "committed"
    assert len(transaction.operations) == 1
    assert transaction.operations[0].status == "completed"


def test_apply_plan_excel_sort_output_differs_when_sort_column_changes(session, tmp_path):
    # Saga #319 wiring-testi konvansiyonu (DESIGN_DECISIONS.md bolum 6) -
    # AYNI PlanStep alaninin (sortColumn) EN AZ IKI farkli degeriyle
    # apply_plan cagirilip GERCEK dosya sisteminde IKI FARKLI gozlemlenebilir
    # sonuc dogrulanmali. Birinci calisma header metniyle ("Tutar"), ikinci
    # calisma bare column-letter ile ("A" - Ad sutunu) sıralar; ciktilarin
    # satir siralamasi birbirinden FARKLI olmali.
    import openpyxl

    run1 = tmp_path / "run1"
    run1.mkdir()
    run2 = tmp_path / "run2"
    run2.mkdir()

    rows = [
        ["Ad", "Tutar"],
        ["Veli", 200],
        ["Ali", 100],
        ["Can", 300],
    ]
    _write_excel(run1, "kaynak.xlsx", rows)
    _write_excel(run2, "kaynak.xlsx", rows)

    pdf_files_1 = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan_1 = _plan([_excel_sort_step(0, "kaynak.xlsx", "Tutar", True, "siralanmis.xlsx")])
    apply_plan(session, plan_1, pdf_files_1, run1)

    pdf_files_2 = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan_2 = _plan([_excel_sort_step(0, "kaynak.xlsx", "A", True, "siralanmis.xlsx")])
    apply_plan(session, plan_2, pdf_files_2, run2)

    wb1 = openpyxl.load_workbook(str(run1 / "siralanmis.xlsx"))
    values_by_tutar = [row[0] for row in wb1.active.iter_rows(min_row=2, values_only=True)]

    wb2 = openpyxl.load_workbook(str(run2 / "siralanmis.xlsx"))
    values_by_ad = [row[0] for row in wb2.active.iter_rows(min_row=2, values_only=True)]

    assert values_by_tutar == ["Ali", "Veli", "Can"]  # 100, 200, 300 sirasinda
    assert values_by_ad == ["Ali", "Can", "Veli"]  # alfabetik Ad sirasinda
    assert values_by_tutar != values_by_ad


def test_apply_plan_excel_sort_resolves_header_text_column(session, tmp_path):
    import openpyxl

    _write_excel(
        tmp_path,
        "kaynak.xlsx",
        [
            ["Ad", "Tutar"],
            ["Veli", 200],
            ["Ali", 100],
        ],
    )
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_sort_step(0, "kaynak.xlsx", "tutar", True, "siralanmis.xlsx")])

    apply_plan(session, plan, pdf_files, tmp_path)

    wb = openpyxl.load_workbook(str(tmp_path / "siralanmis.xlsx"))
    values = [row[1] for row in wb.active.iter_rows(min_row=2, values_only=True)]
    assert values == [100, 200]


def test_apply_plan_excel_sort_resolves_bare_letter_column_when_no_header_matches(session, tmp_path):
    import openpyxl

    _write_excel(
        tmp_path,
        "kaynak.xlsx",
        [
            ["Ad", "Tutar"],
            ["Veli", 200],
            ["Ali", 100],
        ],
    )
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_sort_step(0, "kaynak.xlsx", "B", True, "siralanmis.xlsx")])

    apply_plan(session, plan, pdf_files, tmp_path)

    wb = openpyxl.load_workbook(str(tmp_path / "siralanmis.xlsx"))
    values = [row[1] for row in wb.active.iter_rows(min_row=2, values_only=True)]
    assert values == [100, 200]


def test_apply_plan_rejects_excel_sort_with_unknown_column(session, tmp_path):
    _write_excel(
        tmp_path,
        "kaynak.xlsx",
        [
            ["Ad", "Tutar"],
            ["Veli", 200],
        ],
    )
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_sort_step(0, "kaynak.xlsx", "BilinmeyenSutun", True, "siralanmis.xlsx")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "siralanmis.xlsx").exists()


def test_apply_plan_rejects_excel_sort_of_a_path_outside_allowed_root(session, tmp_path, monkeypatch):
    # REDACT'in ".." teknigiyle AYNI desen (Saga #307) - PdfFileMetadata.filename
    # kendi validator'unda ayrac icermez ama tek segmentlik ".." allowed_root
    # disina cikar. EXCEL_SORT da diger operationType'lar gibi
    # is_path_allowed/validate_plan_paths uzerinden dogrulanmali.
    pdf_files = [PdfFileMetadata(filename="..", createdAt="2026-08-01")]
    plan = _plan([_excel_sort_step(0, "..", "Tutar", True, "siralanmis.xlsx")])

    calls: list = []
    monkeypatch.setattr(
        "backend.orchestrator.excel_sort.sort_excel_sheet",
        lambda *args, **kwargs: calls.append(args),
    )

    with pytest.raises(PathWhitelistError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert calls == []
    assert not (tmp_path / "siralanmis.xlsx").exists()


# Saga #325: EXCEL_FILTER operasyonu (red step) - `OperationType.EXCEL_FILTER`,
# `PlanStep.filterColumn`/`filterValue`/`filteredFileName` ve orchestrator'daki
# EXCEL_FILTER dali henuz YOK. Bu testler simdilik KIRMIZI kalmali
# (AttributeError / ValidationError - EXCEL_SORT'un red-step testleriyle AYNI
# desen).


def _excel_filter_step(
    order: int,
    file_name: str,
    filter_column: str,
    filter_value,
    filtered_file_name: str,
) -> PlanStep:
    return PlanStep(
        order=order,
        operationType=OperationType.EXCEL_FILTER,
        targetFolder="2026-08",
        affectedFileCount=1,
        fileNames=[file_name],
        filterColumn=filter_column,
        filterValue=filter_value,
        filteredFileName=filtered_file_name,
    )


def test_apply_plan_rejects_excel_filter_when_a_data_row_contains_a_formula(session, tmp_path):
    _write_excel_with_formula(
        tmp_path,
        "kaynak.xlsx",
        header=["Ad", "Puan"],
        data_rows=[["Ali", 90], ["Veli", 80]],
        formula_cell="C2",
        formula="=B2+B3",
    )
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_filter_step(0, "kaynak.xlsx", "Puan", 90, "filtrelenmis.xlsx")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "filtrelenmis.xlsx").exists()


def test_apply_plan_filters_excel_rows_when_there_are_no_formulas(session, tmp_path):
    import openpyxl

    _write_excel(
        tmp_path,
        "kaynak.xlsx",
        [
            ["Ad", "Puan"],
            ["Ali", 90],
            ["Veli", 80],
            ["Can", 90],
        ],
    )
    original_bytes = (tmp_path / "kaynak.xlsx").read_bytes()
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_filter_step(0, "kaynak.xlsx", "Puan", 90, "filtrelenmis.xlsx")])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    output_path = tmp_path / "filtrelenmis.xlsx"
    assert output_path.exists()
    wb = openpyxl.load_workbook(str(output_path))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows == [("Ad", "Puan"), ("Ali", 90), ("Can", 90)]
    # Kaynak dosya asla degismemeli.
    assert (tmp_path / "kaynak.xlsx").read_bytes() == original_bytes
    assert transaction.status == "committed"
    assert len(transaction.operations) == 1
    assert transaction.operations[0].status == "completed"


def test_apply_plan_excel_filter_prefers_header_text_over_bare_letter_interpretation(session, tmp_path):
    # Saga #325 P0 senaryosu: "Ad" gibi kisa/alfabetik bir baslik sutun
    # harfi ("AD") ile karistirilmamali - header eslesmesi HER ZAMAN
    # bare-letter yorumundan ONCE denenir.
    import openpyxl

    _write_excel(
        tmp_path,
        "kaynak.xlsx",
        [
            ["Ad", "Puan"],
            ["Ali", 90],
            ["Veli", 80],
        ],
    )
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_filter_step(0, "kaynak.xlsx", "Ad", "Ali", "filtrelenmis.xlsx")])

    apply_plan(session, plan, pdf_files, tmp_path)

    wb = openpyxl.load_workbook(str(tmp_path / "filtrelenmis.xlsx"))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows == [("Ad", "Puan"), ("Ali", 90)]


def test_apply_plan_excel_filter_resolves_bare_letter_column_when_no_header_matches(session, tmp_path):
    import openpyxl

    _write_excel(
        tmp_path,
        "kaynak.xlsx",
        [
            ["Ad", "Puan"],
            ["Ali", 90],
            ["Veli", 80],
        ],
    )
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_filter_step(0, "kaynak.xlsx", "B", 90, "filtrelenmis.xlsx")])

    apply_plan(session, plan, pdf_files, tmp_path)

    wb = openpyxl.load_workbook(str(tmp_path / "filtrelenmis.xlsx"))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows == [("Ad", "Puan"), ("Ali", 90)]


def test_apply_plan_rejects_excel_filter_with_unknown_column(session, tmp_path):
    _write_excel(
        tmp_path,
        "kaynak.xlsx",
        [
            ["Ad", "Puan"],
            ["Veli", 80],
        ],
    )
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_filter_step(0, "kaynak.xlsx", "BilinmeyenSutun", 80, "filtrelenmis.xlsx")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "filtrelenmis.xlsx").exists()


def test_apply_plan_excel_filter_writes_header_only_file_when_no_row_matches(session, tmp_path):
    # ATDD AC-4 / Davranış Sözleşmesi #8: 0 satır eşleşmesi SESSİZ BAŞARI
    # değil, ama hata da değil - header-only dosya yazılır.
    import openpyxl

    _write_excel(
        tmp_path,
        "kaynak.xlsx",
        [
            ["Ad", "Puan"],
            ["Ali", 90],
            ["Veli", 80],
        ],
    )
    pdf_files = [PdfFileMetadata(filename="kaynak.xlsx", createdAt="2026-08-01")]
    plan = _plan([_excel_filter_step(0, "kaynak.xlsx", "Puan", 999, "filtrelenmis.xlsx")])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    wb = openpyxl.load_workbook(str(tmp_path / "filtrelenmis.xlsx"))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows == [("Ad", "Puan")]
    assert transaction.operations[0].status == "completed"


def test_apply_plan_rejects_excel_filter_of_a_path_outside_allowed_root(session, tmp_path, monkeypatch):
    # EXCEL_SORT'un ".." teknigiyle AYNI desen (Saga #307/#324).
    pdf_files = [PdfFileMetadata(filename="..", createdAt="2026-08-01")]
    plan = _plan([_excel_filter_step(0, "..", "Puan", 90, "filtrelenmis.xlsx")])

    calls: list = []
    monkeypatch.setattr(
        "backend.orchestrator.excel_sort.filter_excel_sheet",
        lambda *args, **kwargs: calls.append(args),
    )

    with pytest.raises(PathWhitelistError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert calls == []
    assert not (tmp_path / "filtrelenmis.xlsx").exists()


def test_retry_on_transient_io_error_retries_on_winerror_32(session, tmp_path):
    from backend.orchestrator import _retry_on_transient_io_error

    call_count = {"n": 0}

    def mock_func():
        call_count["n"] += 1
        if call_count["n"] < 2:
            err = OSError("Mock error")
            err.winerror = 32
            raise err
        return "ok"

    with patch("time.sleep"):
        result = _retry_on_transient_io_error(mock_func)
        assert result == "ok"
        assert call_count["n"] == 2


def test_retry_on_transient_io_error_retries_on_winerror_5(session, tmp_path):
    from backend.orchestrator import _retry_on_transient_io_error

    call_count = {"n": 0}

    def mock_func():
        call_count["n"] += 1
        if call_count["n"] < 2:
            err = OSError("Mock error")
            err.winerror = 5
            raise err
        return "ok"

    with patch("time.sleep"):
        result = _retry_on_transient_io_error(mock_func)
        assert result == "ok"
        assert call_count["n"] == 2


def test_retry_on_transient_io_error_fails_after_3_retries(session, tmp_path):
    from backend.orchestrator import _retry_on_transient_io_error

    def mock_func():
        err = OSError("Mock error")
        err.winerror = 32
        raise err

    with patch("time.sleep"):
        with pytest.raises(OSError) as exc_info:
            _retry_on_transient_io_error(mock_func)
        assert exc_info.value.winerror == 32


def test_retry_on_transient_io_error_raises_non_transient_errors(session, tmp_path):
    from backend.orchestrator import _retry_on_transient_io_error

    call_count = {"n": 0}

    def mock_func():
        call_count["n"] += 1
        err = OSError("Mock error")
        err.winerror = 100
        raise err

    with patch("time.sleep") as mock_sleep:
        with pytest.raises(OSError) as exc_info:
            _retry_on_transient_io_error(mock_func)
        assert exc_info.value.winerror == 100
        assert call_count["n"] == 1
        assert mock_sleep.call_count == 0


def test_retry_on_transient_io_error_calls_function_once_without_retries(session, tmp_path):
    from backend.orchestrator import _retry_on_transient_io_error

    def mock_func():
        return "Success"

    with patch("time.sleep") as mock_sleep:
        result = _retry_on_transient_io_error(mock_func)
        assert result == "Success"
        assert mock_sleep.call_count == 0


# Saga #312: purge_oversized_delete_backups — toplam boyut sınırlaması testleri


def _create_backup_files_of_size(backup_dir: Path, size_mb: float) -> None:
    """Create files in backup_dir totaling approximately size_mb megabytes."""
    backup_dir.mkdir(parents=True, exist_ok=True)
    # Create a single file with the specified size
    size_bytes = int(size_mb * 1024 * 1024)
    (backup_dir / "backup.bin").write_bytes(b"x" * size_bytes)


def test_purge_oversized_delete_backups_does_nothing_when_total_size_is_under_threshold(session, tmp_path):
    """Senaryo 1: Toplam boyut < max_total_mb → hiçbir şey silinmez, boş liste"""
    from backend.orchestrator import purge_oversized_delete_backups

    transaction = _apply_a_delete_plan(session, tmp_path, "a.pdf")
    backup_dir = tmp_path / ".windows-ai-files-backup" / str(transaction.id)

    # Create backup files totaling 500 MB (under the 2000 MB threshold)
    _create_backup_files_of_size(backup_dir, 500.0)

    purged_ids = purge_oversized_delete_backups(session, tmp_path, max_total_mb=2000.0)

    assert purged_ids == []
    assert backup_dir.exists()
    assert transaction.status == "committed"


def test_purge_oversized_delete_backups_deletes_oldest_first_until_under_threshold(session, tmp_path):
    """Senaryo 2: Toplam boyut > max_total_mb, 3 aday (eski→yeni) → en eskiden başlanarak sırayla silinir"""
    from backend.orchestrator import purge_oversized_delete_backups

    # Create 3 DELETE-only transactions with different created_at times
    # Transaction 1: oldest, 800 MB
    t1 = _apply_a_delete_plan(session, tmp_path, "old.pdf")
    t1.created_at = dt.datetime.utcnow() - dt.timedelta(days=10)
    backup_dir_1 = tmp_path / ".windows-ai-files-backup" / str(t1.id)
    _create_backup_files_of_size(backup_dir_1, 800.0)
    session.commit()

    # Transaction 2: middle, 700 MB
    t2 = _apply_a_delete_plan(session, tmp_path, "middle.pdf")
    t2.created_at = dt.datetime.utcnow() - dt.timedelta(days=5)
    backup_dir_2 = tmp_path / ".windows-ai-files-backup" / str(t2.id)
    _create_backup_files_of_size(backup_dir_2, 700.0)
    session.commit()

    # Transaction 3: newest, 600 MB
    t3 = _apply_a_delete_plan(session, tmp_path, "new.pdf")
    t3.created_at = dt.datetime.utcnow() - dt.timedelta(days=1)
    backup_dir_3 = tmp_path / ".windows-ai-files-backup" / str(t3.id)
    _create_backup_files_of_size(backup_dir_3, 600.0)
    session.commit()

    # Total: 800 + 700 + 600 = 2100 MB (over 2000 MB threshold)
    # After deleting t1 (800 MB): 1300 MB remaining (under threshold)
    purged_ids = purge_oversized_delete_backups(session, tmp_path, max_total_mb=2000.0)

    # Only the oldest transaction should be purged
    assert t1.id in purged_ids
    assert t2.id not in purged_ids
    assert t3.id not in purged_ids
    assert len(purged_ids) == 1

    # Verify backup directories
    assert not backup_dir_1.exists()  # oldest was deleted
    assert backup_dir_2.exists()  # middle still exists
    assert backup_dir_3.exists()  # newest still exists

    # Verify transaction statuses
    session.expire_all()
    assert session.get(Transaction, t1.id).status == "backup_purged"
    assert session.get(Transaction, t2.id).status == "committed"
    assert session.get(Transaction, t3.id).status == "committed"


def test_purge_oversized_delete_backups_deletes_multiple_when_needed_to_reach_threshold(session, tmp_path):
    """Senaryo 2b: Sınırın altına inmek için birden fazla işlem silinebilir"""
    from backend.orchestrator import purge_oversized_delete_backups

    # Create 3 DELETE-only transactions
    # Transaction 1: oldest, 900 MB
    t1 = _apply_a_delete_plan(session, tmp_path, "old.pdf")
    t1.created_at = dt.datetime.utcnow() - dt.timedelta(days=10)
    backup_dir_1 = tmp_path / ".windows-ai-files-backup" / str(t1.id)
    _create_backup_files_of_size(backup_dir_1, 900.0)
    session.commit()

    # Transaction 2: middle, 800 MB
    t2 = _apply_a_delete_plan(session, tmp_path, "middle.pdf")
    t2.created_at = dt.datetime.utcnow() - dt.timedelta(days=5)
    backup_dir_2 = tmp_path / ".windows-ai-files-backup" / str(t2.id)
    _create_backup_files_of_size(backup_dir_2, 800.0)
    session.commit()

    # Transaction 3: newest, 600 MB
    t3 = _apply_a_delete_plan(session, tmp_path, "new.pdf")
    t3.created_at = dt.datetime.utcnow() - dt.timedelta(days=1)
    backup_dir_3 = tmp_path / ".windows-ai-files-backup" / str(t3.id)
    _create_backup_files_of_size(backup_dir_3, 600.0)
    session.commit()

    # Total: 900 + 800 + 600 = 2300 MB (over 2000 MB threshold)
    # After deleting t1 (900 MB): 1400 MB remaining (under threshold)
    purged_ids = purge_oversized_delete_backups(session, tmp_path, max_total_mb=2000.0)

    assert t1.id in purged_ids
    assert len(purged_ids) == 1
    assert not backup_dir_1.exists()
    assert backup_dir_2.exists()
    assert backup_dir_3.exists()


def test_purge_oversized_delete_backups_stops_as_soon_as_true_recomputed_size_is_under_threshold(session, tmp_path):
    """Senaryo 3: Adayları EN ESKİ'den başlayarak purge et; her purge'dan sonra
    gerçek toplam boyutu yeniden hesapla ve eşik altına inince DUR.

    Phantom/unknown yedeklerin boyutu toplam hesaplamada DAİMA sayılır ama
    purgeable değildir — bu durumda, ilk candidate (t1) purge'dan sonra bile
    toplam (phantom 1500 MB hâlâ var) eşik altında kalır, bu yüzden t2 asla
    dokunulmaz."""
    from backend.orchestrator import purge_oversized_delete_backups

    # Transaction 1: oldest, 500 MB
    t1 = _apply_a_delete_plan(session, tmp_path, "file1.pdf")
    t1.created_at = dt.datetime.utcnow() - dt.timedelta(days=10)
    backup_dir_1 = tmp_path / ".windows-ai-files-backup" / str(t1.id)
    _create_backup_files_of_size(backup_dir_1, 500.0)
    session.commit()

    # Transaction 2: newest, 300 MB
    t2 = _apply_a_delete_plan(session, tmp_path, "file2.pdf")
    t2.created_at = dt.datetime.utcnow() - dt.timedelta(days=1)
    backup_dir_2 = tmp_path / ".windows-ai-files-backup" / str(t2.id)
    _create_backup_files_of_size(backup_dir_2, 300.0)
    session.commit()

    # Create a large "phantom" backup (simulating old backup from another source)
    # that won't be in the candidates list but takes up space
    phantom_backup = tmp_path / ".windows-ai-files-backup" / "9999"
    _create_backup_files_of_size(phantom_backup, 1500.0)

    # Total: 500 + 300 + 1500 = 2300 MB (over 2000 MB threshold)
    # After purging JUST t1 (oldest): 2300 - 500 = 1800 MB (now <= 2000 MB threshold)
    # → function stops, t2 is never purged
    purged_ids = purge_oversized_delete_backups(session, tmp_path, max_total_mb=2000.0)

    # Only t1 should be purged (oldest, and purging it brings us under threshold)
    assert t1.id in purged_ids
    assert t2.id not in purged_ids
    assert len(purged_ids) == 1

    # t1 backup should be deleted, t2 backup should remain
    assert not backup_dir_1.exists()
    assert backup_dir_2.exists()

    # Phantom backup should still exist (it's not a known transaction)
    assert phantom_backup.exists()


def test_purge_oversized_delete_backups_skips_mixed_operation_transactions(session, tmp_path):
    """Senaryo 4: Karışık operasyonlu transaction → hiç dokunulmaz, boyut hesabına dahil olmaz"""
    from backend.orchestrator import purge_oversized_delete_backups

    # Create one pure DELETE transaction (500 MB)
    t_pure = _apply_a_delete_plan(session, tmp_path, "pure_delete.pdf")
    t_pure.created_at = dt.datetime.utcnow() - dt.timedelta(days=10)
    backup_dir_pure = tmp_path / ".windows-ai-files-backup" / str(t_pure.id)
    _create_backup_files_of_size(backup_dir_pure, 500.0)
    session.commit()

    # Create one mixed operation transaction (DELETE + MOVE, 1600 MB)
    # This won't have a backup_dir (or should be skipped)
    _write_pdf(tmp_path, "file1.pdf")
    _write_pdf(tmp_path, "file2.pdf")
    pdf_files = [
        PdfFileMetadata(filename="file1.pdf", createdAt="2026-08-01"),
        PdfFileMetadata(filename="file2.pdf", createdAt="2026-08-02"),
    ]
    plan = _plan(
        [
            _step(0, "2026-08", ["file1.pdf"], operation_type=OperationType.DELETE),
            _step(1, "2026-08", ["file2.pdf"], operation_type=OperationType.MOVE),
        ]
    )
    t_mixed = apply_plan(session, plan, pdf_files, tmp_path)
    t_mixed.created_at = dt.datetime.utcnow() - dt.timedelta(days=5)
    session.commit()

    # Note: mixed operation transaction has no backup_dir, so it won't be
    # counted toward total size or be a purge candidate anyway

    # Total should only count pure DELETE backup: 500 MB (under 2000 MB threshold)
    purged_ids = purge_oversized_delete_backups(session, tmp_path, max_total_mb=2000.0)

    assert purged_ids == []
    assert backup_dir_pure.exists()
    assert t_pure.status == "committed"
    assert t_mixed.status == "committed"


def test_purge_oversized_delete_backups_respects_backup_dir_existence_check(session, tmp_path):
    """Multi-root scenario: candidates where backup_dir doesn't exist are skipped"""
    from backend.orchestrator import purge_oversized_delete_backups

    # Create transaction in tmp_path
    t1 = _apply_a_delete_plan(session, tmp_path, "file1.pdf")
    t1.created_at = dt.datetime.utcnow() - dt.timedelta(days=10)
    backup_dir_1 = tmp_path / ".windows-ai-files-backup" / str(t1.id)
    _create_backup_files_of_size(backup_dir_1, 800.0)
    session.commit()

    # Now delete the backup directory to simulate it belonging to a different root
    shutil.rmtree(backup_dir_1)

    # Call purge with a different root (or after the backup is gone)
    purged_ids = purge_oversized_delete_backups(session, tmp_path, max_total_mb=2000.0)

    # Transaction should not be purged (backup_dir doesn't exist)
    assert t1.id not in purged_ids
    assert t1.status == "committed"


def test_purge_oversized_delete_backups_returns_list_of_purged_transaction_ids(session, tmp_path):
    """Verify that the function returns the list of purged transaction IDs in order"""
    from backend.orchestrator import purge_oversized_delete_backups

    # Create 2 DELETE transactions
    t1 = _apply_a_delete_plan(session, tmp_path, "a.pdf")
    t1.created_at = dt.datetime.utcnow() - dt.timedelta(days=10)
    backup_dir_1 = tmp_path / ".windows-ai-files-backup" / str(t1.id)
    _create_backup_files_of_size(backup_dir_1, 600.0)
    session.commit()

    t2 = _apply_a_delete_plan(session, tmp_path, "b.pdf")
    t2.created_at = dt.datetime.utcnow() - dt.timedelta(days=5)
    backup_dir_2 = tmp_path / ".windows-ai-files-backup" / str(t2.id)
    _create_backup_files_of_size(backup_dir_2, 600.0)
    session.commit()

    # Total: 1200 MB (under 2000 MB, should return empty)
    purged_ids = purge_oversized_delete_backups(session, tmp_path, max_total_mb=2000.0)
    assert purged_ids == []

    # Now call with lower threshold
    purged_ids = purge_oversized_delete_backups(session, tmp_path, max_total_mb=800.0)
    assert t1.id in purged_ids


# Saga #323: APPEND operasyonu (pdf-append-safe, red step) -
# `OperationType.APPEND`, `PlanStep.appendText` ve orchestrator'daki APPEND
# dali henuz YOK. Bu testler simdilik KIRMIZI kalmali (AttributeError /
# ValidationError / ModuleNotFoundError - hepsi beklenen red durumu).
# Referans: artifacts/pdf-append-safe/atdd.md (AC-1..6), plan.md
# (gecici-dosya+atomik-replace deseni, kaynak-yok vs kaynak-bozuk AYRI hata
# mesajlari).


def _append_step(order: int, file_name: str, append_text: str) -> PlanStep:
    return PlanStep(
        order=order,
        operationType=OperationType.APPEND,
        targetFolder="2026-08",
        affectedFileCount=1,
        fileNames=[file_name],
        appendText=append_text,
    )


def test_apply_plan_appends_a_new_page_with_the_given_text_and_commits(session, tmp_path):
    # AC-1: gecerli/okunabilir bir kaynak PDF + appendText verildiginde,
    # sonuc dosyasinda ORIJINAL sayfa sayisi + 1 sayfa olmali.
    from pypdf import PdfReader

    original_page_count = 2
    _write_real_pdf(tmp_path, "rapor.pdf", original_page_count)
    pdf_files = [PdfFileMetadata(filename="rapor.pdf", createdAt="2026-08-01")]
    plan = _plan([_append_step(0, "rapor.pdf", "incelendi ve onaylandı")])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    reader = PdfReader(str(tmp_path / "rapor.pdf"))
    assert len(reader.pages) == original_page_count + 1
    assert transaction.status == "committed"


def test_apply_plan_rejects_append_to_a_corrupt_source_and_leaves_it_untouched(session, tmp_path):
    # AC-2: bozuk bir "PDF" ile APPEND denendiginde PlanApplicationError
    # firlatilmali, mesaj "bozuk"/"okunamiyor" gibi bir ipucu icermeli, VE
    # kaynak dosyanin byte icerigi islem oncesi/sonrasi AYNI kalmali (en
    # kritik assertion - eski projenin kusuru tam olarak bunun kaybolmasiydi).
    corrupt_bytes = b"not a real pdf"
    (tmp_path / "bozuk.pdf").write_bytes(corrupt_bytes)
    pdf_files = [PdfFileMetadata(filename="bozuk.pdf", createdAt="2026-08-01")]
    plan = _plan([_append_step(0, "bozuk.pdf", "incelendi ve onaylandı")])

    with pytest.raises(PlanApplicationError) as excinfo:
        apply_plan(session, plan, pdf_files, tmp_path)

    corrupt_message = str(excinfo.value)
    assert "bozuk" in corrupt_message or "okunamıyor" in corrupt_message or "okunamiyor" in corrupt_message
    assert (tmp_path / "bozuk.pdf").read_bytes() == corrupt_bytes


def test_apply_plan_rejects_append_to_a_missing_source_with_a_distinct_message(session, tmp_path):
    # AC-3: kaynak dosya HIC YOK -> AC-2'nin ("kaynak bozuk") mesajindan
    # AYRI bir "bulunamadi" hatasi donmeli - eski projede bu ikisi AYNI koda
    # dusuyordu, bu task'in cozdugu asil kusur bu.
    corrupt_bytes = b"not a real pdf"
    (tmp_path / "bozuk.pdf").write_bytes(corrupt_bytes)
    pdf_files_corrupt = [PdfFileMetadata(filename="bozuk.pdf", createdAt="2026-08-01")]
    plan_corrupt = _plan([_append_step(0, "bozuk.pdf", "incelendi ve onaylandı")])
    with pytest.raises(PlanApplicationError) as corrupt_excinfo:
        apply_plan(session, plan_corrupt, pdf_files_corrupt, tmp_path)
    corrupt_message = str(corrupt_excinfo.value)

    pdf_files_missing = [PdfFileMetadata(filename="yok.pdf", createdAt="2026-08-01")]
    plan_missing = _plan([_append_step(0, "yok.pdf", "incelendi ve onaylandı")])
    with pytest.raises(PlanApplicationError) as missing_excinfo:
        apply_plan(session, plan_missing, pdf_files_missing, tmp_path)
    missing_message = str(missing_excinfo.value)

    assert "bulunamadı" in missing_message or "bulunamadi" in missing_message
    assert missing_message != corrupt_message


@pytest.mark.skipif(
    os.name == "nt",
    reason=(
        "Windows NTFS'te os.chmod(path, 0o000)/read-only bayragi dosya "
        "sahibinin okuma/yazmasini AYNI sekilde engellemiyor (bkz. "
        "test_file_search.py'deki ayni desen) - bu test Unix-only."
    ),
)
def test_apply_plan_rejects_append_to_a_permission_denied_source(session, tmp_path):
    # AC-4: salt-okunur/izin hatasi olan bir kaynakla APPEND denemesi hata
    # dondurmeli, dosyaya dokunulmamali.
    _write_real_pdf(tmp_path, "kilitli.pdf", 1)
    target = tmp_path / "kilitli.pdf"
    original_bytes = target.read_bytes()
    os.chmod(target, 0o000)
    try:
        pdf_files = [PdfFileMetadata(filename="kilitli.pdf", createdAt="2026-08-01")]
        plan = _plan([_append_step(0, "kilitli.pdf", "incelendi ve onaylandı")])

        with pytest.raises(PlanApplicationError):
            apply_plan(session, plan, pdf_files, tmp_path)
    finally:
        os.chmod(target, 0o644)

    assert target.read_bytes() == original_bytes


def test_append_text_empty_or_whitespace_only_is_rejected_by_pydantic():
    # AC-5: appendText bos/whitespace-only ise Pydantic ValidationError
    # firlatmali - model seviyesinde, apply_plan'a hic ulasmadan.
    with pytest.raises(ValidationError):
        _append_step(0, "rapor.pdf", "")

    with pytest.raises(ValidationError):
        _append_step(0, "rapor.pdf", "   ")


def test_apply_plan_rejects_append_of_a_path_outside_allowed_root(session, tmp_path):
    # AC-6: whitelist disi bir hedef ile APPEND plani - mevcut MERGE/REDACT
    # testlerindeki whitelist-ihlali testinin AYNISI (bkz.
    # test_apply_plan_rejects_redact_of_a_path_outside_allowed_root).
    pdf_files = [PdfFileMetadata(filename="..", createdAt="2026-08-01")]
    plan = _plan([_append_step(0, "..", "incelendi ve onaylandı")])

    with pytest.raises(PathWhitelistError):
        apply_plan(session, plan, pdf_files, tmp_path)


# Saga #321: pdf-sayfa-araligi-secimi (red step) -
# `OperationType.PDF_EXTRACT_PAGES`/`OperationType.PDF_DELETE_PAGES`,
# `PlanStep.pageSpec`/`extractedFileName`/`remainingFileName` ve
# orchestrator'daki iki step-uygulama dali henuz YOK. Bu testler simdilik
# KIRMIZI kalmali (AttributeError / ValidationError - EXCEL_FILTER'in
# red-step testleriyle AYNI desen).


def _write_tagged_pdf(root, filename: str, page_count: int) -> None:
    # _write_real_pdf'in tum sayfalari ayni boyutta uretmesi, secili
    # sayfa KIMLIGININ dogrulanmasini imkansiz kilar (red-team bulgu 1,
    # Saga #321) - bu yardimci her sayfaya ayirt edici bir genislik
    # (100+i, i=0-indexed sayfa sirasi) verir.
    from pypdf import PdfWriter

    writer = PdfWriter()
    for i in range(page_count):
        writer.add_blank_page(width=100 + i, height=100 + i)
    with open(root / filename, "wb") as f:
        writer.write(f)
    writer.close()


def _pdf_extract_pages_step(
    order: int,
    file_name: str,
    page_spec: str,
    extracted_file_name: str,
) -> PlanStep:
    return PlanStep(
        order=order,
        operationType=OperationType.PDF_EXTRACT_PAGES,
        targetFolder="2026-08",
        affectedFileCount=1,
        fileNames=[file_name],
        pageSpec=page_spec,
        extractedFileName=extracted_file_name,
    )


def _pdf_delete_pages_step(
    order: int,
    file_name: str,
    page_spec: str,
    remaining_file_name: str,
) -> PlanStep:
    return PlanStep(
        order=order,
        operationType=OperationType.PDF_DELETE_PAGES,
        targetFolder="2026-08",
        affectedFileCount=1,
        fileNames=[file_name],
        pageSpec=page_spec,
        remainingFileName=remaining_file_name,
    )


def test_apply_plan_extracts_pdf_pages_happy_path(session, tmp_path):
    from pypdf import PdfReader

    _write_tagged_pdf(tmp_path, "kaynak.pdf", 10)
    original_bytes = (tmp_path / "kaynak.pdf").read_bytes()
    pdf_files = [PdfFileMetadata(filename="kaynak.pdf", createdAt="2026-08-01")]
    plan = _plan([_pdf_extract_pages_step(0, "kaynak.pdf", "1,3,5-9", "cikarilan.pdf")])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    output_path = tmp_path / "cikarilan.pdf"
    assert output_path.exists()
    reader = PdfReader(str(output_path))
    assert len(reader.pages) == 7
    # Sadece SAYIYI degil, HANGI sayfalarin ORIJINAL sirayla secildigini de
    # dogrula (red-team bulgu 1, Saga #321).
    assert [p.mediabox.width for p in reader.pages] == [100, 102, 104, 105, 106, 107, 108]
    # Kaynak dosya asla degismemeli.
    assert (tmp_path / "kaynak.pdf").read_bytes() == original_bytes
    assert transaction.status == "committed"
    assert len(transaction.operations) == 1
    assert transaction.operations[0].status == "completed"


def test_apply_plan_deletes_pdf_pages_happy_path(session, tmp_path):
    from pypdf import PdfReader

    _write_tagged_pdf(tmp_path, "kaynak.pdf", 10)
    original_bytes = (tmp_path / "kaynak.pdf").read_bytes()
    pdf_files = [PdfFileMetadata(filename="kaynak.pdf", createdAt="2026-08-01")]
    plan = _plan([_pdf_delete_pages_step(0, "kaynak.pdf", "1,3,5-9", "kalan.pdf")])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    output_path = tmp_path / "kalan.pdf"
    assert output_path.exists()
    reader = PdfReader(str(output_path))
    # 10 sayfadan [1,3,5,6,7,8,9] silinince [2,4,10] kalir -> 3 sayfa.
    assert len(reader.pages) == 3
    # Sadece SAYIYI degil, kalan sayfalarin DOGRU KIMLIKTE ve ORIJINAL
    # sirada oldugunu da dogrula (red-team bulgu 1, Saga #321).
    assert [p.mediabox.width for p in reader.pages] == [101, 103, 109]
    # Kaynak dosya asla degismemeli.
    assert (tmp_path / "kaynak.pdf").read_bytes() == original_bytes
    assert transaction.status == "committed"
    assert len(transaction.operations) == 1
    assert transaction.operations[0].status == "completed"


def test_apply_plan_rejects_pdf_extract_pages_with_reversed_range(session, tmp_path):
    _write_real_pdf(tmp_path, "kaynak.pdf", 10)
    pdf_files = [PdfFileMetadata(filename="kaynak.pdf", createdAt="2026-08-01")]
    plan = _plan([_pdf_extract_pages_step(0, "kaynak.pdf", "9-5", "cikarilan.pdf")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "cikarilan.pdf").exists()


def test_apply_plan_rejects_pdf_extract_pages_with_out_of_document_page_number(session, tmp_path):
    _write_real_pdf(tmp_path, "kaynak.pdf", 5)
    pdf_files = [PdfFileMetadata(filename="kaynak.pdf", createdAt="2026-08-01")]
    plan = _plan([_pdf_extract_pages_step(0, "kaynak.pdf", "1,3,8", "cikarilan.pdf")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "cikarilan.pdf").exists()


def test_apply_plan_rejects_pdf_delete_pages_with_reversed_range(session, tmp_path):
    _write_real_pdf(tmp_path, "kaynak.pdf", 10)
    pdf_files = [PdfFileMetadata(filename="kaynak.pdf", createdAt="2026-08-01")]
    plan = _plan([_pdf_delete_pages_step(0, "kaynak.pdf", "9-5", "kalan.pdf")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "kalan.pdf").exists()


def test_apply_plan_rejects_pdf_delete_pages_with_out_of_document_page_number(session, tmp_path):
    _write_real_pdf(tmp_path, "kaynak.pdf", 5)
    pdf_files = [PdfFileMetadata(filename="kaynak.pdf", createdAt="2026-08-01")]
    plan = _plan([_pdf_delete_pages_step(0, "kaynak.pdf", "1,3,8", "kalan.pdf")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "kalan.pdf").exists()


def test_apply_plan_rejects_pdf_delete_pages_when_all_pages_are_deleted(session, tmp_path):
    # ATDD AC-5: tum sayfalar silinirse (0 sayfa kalir) PlanApplicationError,
    # hicbir dosya yazilmaz - EXCEL_FILTER'in "0 satir eslesti = basari"
    # kararindan KASITLI OLARAK FARKLI.
    _write_real_pdf(tmp_path, "kaynak.pdf", 3)
    pdf_files = [PdfFileMetadata(filename="kaynak.pdf", createdAt="2026-08-01")]
    plan = _plan([_pdf_delete_pages_step(0, "kaynak.pdf", "1-3", "kalan.pdf")])

    with pytest.raises(PlanApplicationError, match="tüm sayfaları silinemez"):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "kalan.pdf").exists()


def test_apply_plan_rejects_pdf_extract_pages_of_a_path_outside_allowed_root(session, tmp_path, monkeypatch):
    # EXCEL_FILTER'in ".." teknigiyle AYNI desen (Saga #307/#324/#325).
    pdf_files = [PdfFileMetadata(filename="..", createdAt="2026-08-01")]
    plan = _plan([_pdf_extract_pages_step(0, "..", "1,3", "cikarilan.pdf")])

    calls: list = []
    monkeypatch.setattr(
        "backend.orchestrator.pdf_pages.extract_pdf_pages",
        lambda *args, **kwargs: calls.append(args),
    )

    with pytest.raises(PathWhitelistError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert calls == []
    assert not (tmp_path / "cikarilan.pdf").exists()


def test_apply_plan_rejects_pdf_delete_pages_of_a_path_outside_allowed_root(session, tmp_path, monkeypatch):
    # EXCEL_FILTER'in ".." teknigiyle AYNI desen (Saga #307/#324/#325).
    pdf_files = [PdfFileMetadata(filename="..", createdAt="2026-08-01")]
    plan = _plan([_pdf_delete_pages_step(0, "..", "1,3", "kalan.pdf")])

    calls: list = []
    monkeypatch.setattr(
        "backend.orchestrator.pdf_pages.delete_pdf_pages",
        lambda *args, **kwargs: calls.append(args),
    )

    with pytest.raises(PathWhitelistError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert calls == []
    assert not (tmp_path / "kalan.pdf").exists()


# Saga #322: pdf-sikistirma (red step) - `OperationType.PDF_COMPRESS`,
# `PlanStep.compressedFileName` ve orchestrator'daki step-uygulama dali
# henuz YOK. Bu testler simdilik KIRMIZI kalmali (AttributeError /
# ValidationError - EXCEL_FILTER/PDF_EXTRACT_PAGES'in red-step testleriyle
# AYNI desen).
#
# plan.md'nin tasarim karari: `compress_pdf()` `True` donerse EXCEL_FILTER
# bloğunun (record_file_operation + status="completed" + applied.append)
# kopyasi uygulanir; `False` donerse (buyume korumasi) HICBIR
# `FileOperation` kaydi olusturulmadan `continue` (LIST'in izlediği desen)
# - bu testler tam olarak bu ayrimi dogrular.


def _write_compressible_pdf_for_orchestrator(root, filename: str, page_count: int = 5) -> None:
    # test_pdf_compress.py'deki _write_compressible_pdf ile AYNI teknik:
    # reportlab, pageCompression=0 ile SIKISTIRILMAMIS, tekrarlanan buyuk
    # content stream'leri olan sayfalar uretir - pypdf'in
    # compress_content_streams() (flate encode) bunu belirgin sekilde
    # kucultmeli.
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(root / filename), pageCompression=0)
    repeated_line = "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA "
    for _ in range(page_count):
        y = 780
        for _ in range(60):
            c.drawString(20, y, repeated_line)
            y -= 10
            if y < 20:
                break
        c.showPage()
    c.save()


def _pdf_compress_step(order: int, file_name: str, compressed_file_name: str) -> PlanStep:
    return PlanStep(
        order=order,
        operationType=OperationType.PDF_COMPRESS,
        targetFolder="2026-08",
        affectedFileCount=1,
        fileNames=[file_name],
        compressedFileName=compressed_file_name,
    )


def test_apply_plan_compresses_pdf_happy_path(session, tmp_path):
    _write_compressible_pdf_for_orchestrator(tmp_path, "kaynak.pdf")
    original_bytes = (tmp_path / "kaynak.pdf").read_bytes()
    pdf_files = [PdfFileMetadata(filename="kaynak.pdf", createdAt="2026-08-01")]
    plan = _plan([_pdf_compress_step(0, "kaynak.pdf", "sikistirilmis.pdf")])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    output_path = tmp_path / "sikistirilmis.pdf"
    assert output_path.exists()
    assert output_path.stat().st_size < len(original_bytes)
    # Kaynak dosya asla degismemeli.
    assert (tmp_path / "kaynak.pdf").read_bytes() == original_bytes
    assert transaction.status == "committed"
    assert len(transaction.operations) == 1
    assert transaction.operations[0].status == "completed"


def test_apply_plan_skips_recording_an_operation_when_pdf_compression_does_not_shrink_the_file(session, tmp_path):
    # ATDD AC-2/AC-5: buyume korumasi - sikisma saglanamazsa (zaten
    # minimal PDF) HICBIR FileOperation kaydi olusmaz (LIST deseni),
    # compressedFileName YAZILMAZ, ama transaction yine de "committed"
    # (hata DEGIL).
    _write_real_pdf(tmp_path, "kaynak.pdf", 1)
    original_bytes = (tmp_path / "kaynak.pdf").read_bytes()
    pdf_files = [PdfFileMetadata(filename="kaynak.pdf", createdAt="2026-08-01")]
    plan = _plan([_pdf_compress_step(0, "kaynak.pdf", "sikistirilmis.pdf")])

    transaction = apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "sikistirilmis.pdf").exists()
    assert (tmp_path / "kaynak.pdf").read_bytes() == original_bytes
    assert transaction.status == "committed"
    assert len(transaction.operations) == 0


def test_apply_plan_rejects_pdf_compress_of_a_broken_source(session, tmp_path):
    # AC-3: kaynak PDF bozuk/pypdf acamiyor -> PlanApplicationError,
    # hicbir dosya yazilmaz.
    (tmp_path / "bozuk.pdf").write_bytes(b"this is not a pdf at all")
    pdf_files = [PdfFileMetadata(filename="bozuk.pdf", createdAt="2026-08-01")]
    plan = _plan([_pdf_compress_step(0, "bozuk.pdf", "sikistirilmis.pdf")])

    with pytest.raises(PlanApplicationError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert not (tmp_path / "sikistirilmis.pdf").exists()


def test_apply_plan_rejects_pdf_compress_of_a_path_outside_allowed_root(session, tmp_path, monkeypatch):
    # EXCEL_FILTER/PDF_EXTRACT_PAGES'in ".." teknigiyle AYNI desen
    # (Saga #307/#324/#325/#321).
    pdf_files = [PdfFileMetadata(filename="..", createdAt="2026-08-01")]
    plan = _plan([_pdf_compress_step(0, "..", "sikistirilmis.pdf")])

    calls: list = []
    monkeypatch.setattr(
        "backend.orchestrator.pdf_compress.compress_pdf",
        lambda *args, **kwargs: calls.append(args),
    )

    with pytest.raises(PathWhitelistError):
        apply_plan(session, plan, pdf_files, tmp_path)

    assert calls == []
    assert not (tmp_path / "sikistirilmis.pdf").exists()
