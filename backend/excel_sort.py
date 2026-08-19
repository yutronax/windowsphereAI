import os
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string


class ExcelSortFormulaGuardError(Exception):
    """Sıralanacak veri aralığında en az bir GERÇEK formül hücresi
    (`cell.data_type == "f"`) bulunduğunda fırlatılır (Saga #324, ATDD
    P0 #1) - openpyxl fiziksel satır taşımada formül metnini otomatik
    güncellemediği için, formüllü bir aralığı sessizce sıralamak veriyi
    sessizce bozar. Bu yüzden işlem TAMAMEN reddedilir, sıfır satır
    değiştirilir, çıktı dosyası hiç oluşturulmaz."""


class ExcelFilterFormulaGuardError(Exception):
    """Filtrelenecek veri aralığında en az bir GERÇEK formül hücresi
    (`cell.data_type == "f"`) bulunduğunda fırlatılır (Saga #325, ATDD
    AC-5) - ExcelSortFormulaGuardError ile AYNI gerekçe: satır alt
    kümesi yeni dosyaya kopyalanırken hücre konumları değişebilir,
    formül metni otomatik güncellenmez. İşlem TAMAMEN reddedilir, sıfır
    satır işlenir, çıktı dosyası hiç oluşturulmaz."""


def resolve_sort_column(header_row: tuple, column_spec: str) -> int:
    """`column_spec`i 0-indexed bir sütun numarasına çözer (Saga #324,
    ATDD P0 #3 / P1 #4/#5 - `header_row` tuple'ının Python indexiyle
    doğrudan uyumlu, bkz. test_excel_sort.py sözleşmesi). Önce
    `header_row`daki (case-insensitive, kırpılmış) metin eşleşmesi
    denenir - bir eşleşme varsa bare column-letter yorumundan ÖNCE
    kazanır. Eşleşme yoksa `column_spec` geçerli bir Excel sütun harfi
    (ör. "C", "AA") olarak yorumlanmaya çalışılır. İkisi de başarısız
    olursa `ValueError` fırlatılır."""
    normalized_spec = column_spec.strip().lower()
    if normalized_spec == "":
        raise ValueError("column_spec must not be blank")

    for index, cell_value in enumerate(header_row):
        if isinstance(cell_value, str) and cell_value.strip().lower() == normalized_spec:
            return index

    try:
        return column_index_from_string(column_spec.strip().upper()) - 1
    except ValueError:
        raise ValueError(
            f"Sütun ne header metniyle ne geçerli bir sütun harfiyle çözülebildi: '{column_spec}'"
        ) from None


def sort_excel_sheet(source_path: Path, sort_column: str, ascending: bool, destination_path: Path) -> None:
    """`source_path`teki .xlsx dosyasının aktif sayfasını `sort_column`a
    göre (artan/azalan) sıralar ve sonucu YENİ bir dosyaya
    (`destination_path`) yazar - kaynağa asla dokunulmaz (Saga #324, ATDD
    P0 #2). İlk satır header sayılır, 2..max_row veri satırlarıdır.

    Veri aralığında herhangi bir hücre GERÇEK bir formül ise
    (`cell.data_type == "f"`) `ExcelSortFormulaGuardError` fırlatılır,
    hiçbir dosya yazılmaz (ATDD P0 #1). 0 veri satırı varsa (boş sayfa
    veya sadece header) işlem no-op sayılır - dosya olduğu gibi
    `destination_path`'e kopyalanır, hata FIRLATILMAZ (ATDD P2 #6).

    `_forward_merge`/`_forward_redact` (orchestrator.py) ile AYNI
    geçici-dosya-yaz + atomik `Path.replace` deseni kullanılır - yazma
    yarıda kesilirse `destination_path`'te YARIM/BOZUK bir dosya
    KALMAZ. Path whitelist kontrolü BU MODÜLDE YAPILMAZ - merkezi
    doğrulama backend/security.py'de yapılır (Saga #307/#319
    konvansiyonu, DESIGN_DECISIONS.md §6)."""
    workbook = openpyxl.load_workbook(str(source_path))
    sheet = workbook.active

    max_row = sheet.max_row
    data_row_count = max(max_row - 1, 0) if max_row and max_row >= 1 else 0

    fd, temp_path_str = tempfile.mkstemp(
        suffix=".tmp", prefix=destination_path.name + ".", dir=str(destination_path.parent)
    )
    temp_path = Path(temp_path_str)
    os.close(fd)
    try:
        if data_row_count <= 0:
            # ATDD P2 #6: sıralanacak veri satırı yok - no-op, dosya
            # olduğu gibi kopyalanır, sütun çözümlemesi/formül taraması
            # anlamsız olduğu için hiç YAPILMAZ.
            workbook.save(str(temp_path))
            temp_path.replace(destination_path)
            return

        header_row = next(iter(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        sort_column_index = resolve_sort_column(header_row, sort_column)

        for row in sheet.iter_rows(min_row=2, max_row=max_row):
            for cell in row:
                if cell.data_type == "f":
                    raise ExcelSortFormulaGuardError(
                        f"Sıralanacak veri aralığında formül bulundu ('{cell.coordinate}'), "
                        "işlem güvenle yapılamıyor - reddedildi."
                    )

        data_rows = [
            list(row) for row in sheet.iter_rows(min_row=2, max_row=max_row, values_only=True)
        ]
        data_rows.sort(key=lambda row: _sort_key(row, sort_column_index), reverse=not ascending)

        for row_offset, row_values in enumerate(data_rows):
            row_number = row_offset + 2
            for column_offset, value in enumerate(row_values):
                sheet.cell(row=row_number, column=column_offset + 1, value=value)

        workbook.save(str(temp_path))
        temp_path.replace(destination_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _sort_key(row: list, sort_index: int):
    value = row[sort_index] if sort_index < len(row) else None
    # None değerler ile karşılaştırılabilir olmayan tipler (str vs int)
    # arasındaki sıralama karşılaştırmasını patlatmamak için, None'ı her
    # zaman en küçük kabul eden bir (has_value, value) anahtarı kullanılır.
    return (value is None, value if value is not None else "")


def filter_excel_sheet(
    source_path: Path, filter_column: str, filter_value, destination_path: Path
) -> None:
    """`source_path`teki .xlsx dosyasının aktif sayfasında `filter_column`
    (header metni veya sütun harfi, `resolve_sort_column` ile çözülür)
    değeri `filter_value`ya EŞİT olan satırları (+ header) YENİ bir
    dosyaya (`destination_path`) yazar - kaynağa asla dokunulmaz (Saga
    #325, ATDD Davranış Sözleşmesi #1, `sort_excel_sheet` ile AYNI desen).

    Karşılaştırma `str(hücre) == str(filter_value)` ile yapılır. Hiçbir
    satır eşleşmezse hata FIRLATILMAZ - sadece header içeren bir dosya
    yazılır (ATDD AC-4, sessiz başarı değil ama hata da değil).

    Veri aralığında herhangi bir hücre GERÇEK bir formül ise
    (`cell.data_type == "f"`) `ExcelFilterFormulaGuardError` fırlatılır,
    hiçbir dosya yazılmaz (ATDD AC-5). 0 veri satırı varsa (boş sayfa
    veya sadece header) işlem no-op sayılır - dosya olduğu gibi
    `destination_path`'e kopyalanır, hata FIRLATILMAZ (ATDD AC-6).

    `sort_excel_sheet` ile AYNI geçici-dosya-yaz + atomik `Path.replace`
    deseni kullanılır - yazma yarıda kesilirse `destination_path`'te
    YARIM/BOZUK bir dosya KALMAZ. Path whitelist kontrolü BU MODÜLDE
    YAPILMAZ - merkezi doğrulama backend/security.py'de yapılır."""
    workbook = openpyxl.load_workbook(str(source_path))
    sheet = workbook.active

    max_row = sheet.max_row
    data_row_count = max(max_row - 1, 0) if max_row and max_row >= 1 else 0

    fd, temp_path_str = tempfile.mkstemp(
        suffix=".tmp", prefix=destination_path.name + ".", dir=str(destination_path.parent)
    )
    temp_path = Path(temp_path_str)
    os.close(fd)
    try:
        if data_row_count <= 0:
            # ATDD AC-6: filtrelenecek veri satırı yok - no-op, dosya
            # olduğu gibi kopyalanır, sütun çözümlemesi/formül taraması
            # anlamsız olduğu için hiç YAPILMAZ.
            workbook.save(str(temp_path))
            temp_path.replace(destination_path)
            return

        header_row = next(iter(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        filter_column_index = resolve_sort_column(header_row, filter_column)

        for row in sheet.iter_rows(min_row=2, max_row=max_row):
            for cell in row:
                if cell.data_type == "f":
                    raise ExcelFilterFormulaGuardError(
                        f"Filtrelenecek veri aralığında formül bulundu ('{cell.coordinate}'), "
                        "işlem güvenle yapılamıyor - reddedildi."
                    )

        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.append(list(header_row))
        target_value = str(filter_value)
        for row in sheet.iter_rows(min_row=2, max_row=max_row, values_only=True):
            cell_value = row[filter_column_index] if filter_column_index < len(row) else None
            if str(cell_value) == target_value:
                new_sheet.append(list(row))

        new_workbook.save(str(temp_path))
        temp_path.replace(destination_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
